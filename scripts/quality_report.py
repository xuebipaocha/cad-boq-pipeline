# -*- coding: utf-8 -*-
"""质量聚合报告 — v5.14 工作流 P2

把 审图记录.json / 算量质量报告.json / 清单质量报告.json 聚合成一张总表,
用户一次看全三个环节的质量状态。

用法:
  python3 quality_report.py --output-dir output
  或在 pipeline 全流程末尾自动调用(见 pipeline.py)。

输出: output/质量聚合报告.json + 质量聚合报告.xlsx
"""
import json
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def _load(output_dir, name):
    p = os.path.join(output_dir, name)
    if not os.path.exists(p):
        return None
    try:
        with open(p, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def aggregate(output_dir):
    """聚合三份质量报告 → 总表 dict。缺失环节标记 '未执行'。"""
    review = _load(output_dir, '审图记录.json')
    calc = _load(output_dir, '算量质量报告.json')
    boq = _load(output_dir, '清单质量报告.json')

    def _warn_count(r, key='警告'):
        if not r:
            return None
        return len(r.get(key, []) or [])

    def _top_warns(r, n=3):
        if not r:
            return []
        ws = r.get('警告', []) or []
        return [{'级别': w.get('级别', ''), '问题': w.get('问题', '')} for w in ws[:n]]

    report = {
        '生成时间': None,  # 由 run() 填充
        '环节': {
            '审图': {
                '状态': '完成' if review else '未执行',
                '问题数': review.get('total', len(review.get('problems', []))) if review else None,
                '警告数': _warn_count(review),
                '主要问题': _top_warns(review),
            },
            '算量': {
                '状态': '完成' if calc else '未执行',
                '质量分': calc.get('质量分') if calc else None,
                '警告数': _warn_count(calc),
                '主要问题': _top_warns(calc),
            },
            '编清单': {
                '状态': '完成' if boq else '未执行',
                '质量分': boq.get('质量分') if boq else None,
                '匹配率': boq.get('匹配率') if boq else None,
                '高置信度': boq.get('高置信度') if boq else None,
                '低置信度': boq.get('低置信度') if boq else None,
                '待匹配': boq.get('待匹配') if boq else None,
                '估算值': boq.get('估算值') if boq else None,
                '待提取': boq.get('待提取') if boq else None,
                '警告数': _warn_count(boq),
                '主要问题': _top_warns(boq),
            },
        },
    }
    # 总评: 三环节质量分加权(缺失环节不计)
    scores = [calc.get('质量分') for calc in [calc] if calc and calc.get('质量分') is not None]
    if boq and boq.get('质量分') is not None:
        scores.append(boq['质量分'])
    if review is not None:
        # 审图无质量分: 问题多则扣分(0 问题=100, 每 5 个问题扣 5 分, 下限 40)
        n_probs = review.get('total', 0)
        scores.append(max(40, 100 - (n_probs // 5) * 5))
    if scores:
        report['总质量分'] = round(sum(scores) / len(scores), 1)
    else:
        report['总质量分'] = None
    return report


def export_excel(report, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = '质量聚合报告'
    hf = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hft = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    df = Font(name='微软雅黑', size=10)
    bd = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=1, column=1, value='CAD-BOQ 质量聚合报告').font = Font(name='微软雅黑', bold=True, size=14)
    ws.cell(row=2, column=1, value=f"总质量分: {report.get('总质量分', 'N/A')}").font = Font(name='微软雅黑', size=10)
    headers = ['环节', '状态', '质量分', '问题/警告数', '关键指标', '主要问题']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = hft
        c.fill = hf
        c.alignment = Alignment(horizontal='center')
        c.border = bd
    r = 5
    for stage, info in report.get('环节', {}).items():
        if stage == '审图':
            key = f"问题数: {info.get('问题数')}"
        elif stage == '算量':
            key = f"警告数: {info.get('警告数')}"
        else:
            key = f"匹配率: {info.get('匹配率')} | 待匹配: {info.get('待匹配')} | 低置信度: {info.get('低置信度')}"
        probs = '; '.join(f"[{w['级别']}]{w['问题']}" for w in info.get('主要问题', []))[:120]
        vals = [stage, info.get('状态'), info.get('质量分', ''), info.get('警告数', ''),
                key, probs]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = df
            c.border = bd
            c.alignment = Alignment(vertical='center', wrap_text=True)
        r += 1
    for c, w in [('A', 10), ('B', 8), ('C', 8), ('D', 12), ('E', 32), ('F', 60)]:
        ws.column_dimensions[c].width = w
    wb.save(path)


def run(output_dir):
    report = aggregate(output_dir)
    import datetime
    report['生成时间'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    json_path = os.path.join(output_dir, '质量聚合报告.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    try:
        xlsx_path = os.path.join(output_dir, '质量聚合报告.xlsx')
        export_excel(report, xlsx_path)
    except Exception as e:
        print(f'  ⚠ 质量聚合 xlsx 导出失败: {e}')
        xlsx_path = None
    print(f'  质量聚合: 总质量分 {report.get("总质量分")} | 审图{report["环节"]["审图"]["状态"]} '
          f'算量{report["环节"]["算量"]["状态"]} 编清单{report["环节"]["编清单"]["状态"]}')
    print(f'  输出: {json_path}')
    return report


if __name__ == '__main__':
    import argparse
    ap = argparse.ArgumentParser(description='质量聚合报告: 审图/算量/清单三质量合一')
    ap.add_argument('--output-dir', default=os.path.join(os.path.dirname(BASE_DIR), 'output'),
                    help='输出目录(含三个质量 json)')
    args = ap.parse_args()
    run(args.output_dir)
