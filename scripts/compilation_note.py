# -*- coding: utf-8 -*-
"""编制说明自动生成 — v6.7

专业造价工程师交付成果的必备部分: 编制依据/工程概况/编制口径/遗留问题/主要指标。
当前工作流此前完全没有编制说明(相当于交了答案没交过程)。

输出: 编制说明.xlsx(独立成果, 与计算书/清单配套交付)。
"""
import os
import json


def build_note(pid, calc_results, boq_items, pending_items, estimated_items,
               problems=None, quotas_note='', project_name=''):
    """生成编制说明结构化内容。

    pid: 识图结果 | calc_results: 算量结果 | boq_items: 清单结果
    pending_items: 待提取清单 | estimated_items: 待核实清单
    problems: 审图问题列表 | project_name: 工程名称(step4 已提取)
    """
    meta = pid.get('图纸元数据', {}) or {}
    areas = pid.get('面积区域', []) or []
    bfa = max((float(a.get('面积_m2', 0) or 0) for a in areas), default=0)
    perim = max((float(a.get('周长_m', 0) or 0) for a in areas), default=0)
    area_src = next((a.get('面积来源', '') for a in areas if a.get('面积_m2') == bfa), '')
    dn = pid.get('设计说明', {}) or {}
    prof = dn.get('工程概况', {}) or {}
    proj = project_name or pid.get('工程名称') or meta.get('图名') or '未识别'
    sections = {}

    # 一、编制依据
    basis = [
        f"1. 图纸: {proj}"
        f"{('，图号 ' + str(meta.get('图号', ''))) if meta.get('图号') else ''}",
        "2. 计量规范: 《建设工程工程量清单计价规范》GB50500-2013"
        "、《房屋建筑与装饰工程工程量计算规范》GB50854-2013",
        "3. 计价定额: 《辽宁省建设工程计价依据(2024)》(辽住建发)",
        "4. 图纸会审记录及设计变更(如有)",
    ]
    if quotas_note:
        basis.append(f"5. {quotas_note}")
    sections['编制依据'] = basis

    # 二、工程概况
    overview = [
        f"1. 工程名称: {proj}",
        f"2. 专业类别: {pid.get('专业类型', '未识别')}",
        f"3. 工程性质: {pid.get('工程性质', '新建')}",
        f"4. 建筑面积: {bfa:.2f}m²（来源: {area_src or '未识别'}）" if bfa else "4. 建筑面积: 未识别",
    ]
    ver = pid.get('图纸版本', {}) or {}
    if ver.get('修订记录'):
        overview.append(f"5. 图纸版本: 修订记录 {'、'.join(ver['修订记录'])}（以最新修订为准）")
    if perim:
        overview.append(f"5. 主区域周长: {perim:.2f}m")
    for k, v in prof.items():
        if v and k not in ('层数',):
            overview.append(f"6. 概况({k}): {v}")
    layers_n = len(pid.get('构造层', []) or [])
    rooms = pid.get('房间', []) or []
    rooms_n = sum(int(r.get('数量', 1) or 1) for r in rooms)
    overview.append(f"7. 识图概况: 构造层{layers_n}层、房间{rooms_n}间、"
                    f"门窗{len(pid.get('门窗', []) or [])}种、"
                    f"钢构件{len((pid.get('钢结构') or {}).get('构件', []) or [])}个")
    sections['工程概况'] = overview

    # 三、编制口径
    scope = [
        "1. 工程量按图纸所示尺寸以 GB50854-2013 计算规则计算；",
        "2. 面积口径: 以图签/闭合轮廓识别的建筑面积为准，做法分劈按房间几何实测，"
        "无几何证据的部位保持估算并标'待核'（不编造分劈数）；",
        "3. 数据来源四级标记: 实测/文字标注/估算/待提取，其中估算项已分流'待核实清单'，"
        "待提取项已分流'待提取清单'，均不混入正式清单；",
        "4. 混凝土构件按图示截面尺寸计算，梁柱重叠、板扣洞口、墙扣门窗已计入；",
        "5. 钢筋按图纸配筋信息，无配筋信息时按含钢量估算并标记。",
    ]
    sections['编制口径'] = scope

    # 四、遗留问题(图纸疑问/待核实/待提取) — v6.9: 每条带影响分项(图纸会审形态)
    calc_names = [i.get('source_name', '') for i in boq_items] + \
                 [e.get('分项名称', '') for e in (estimated_items or [])] + \
                 [p.get('分项名称', '') for p in (pending_items or [])]

    def _link(qtext):
        """问题文本关键词 → 影响的算量/清单分项(前3个)。"""
        if not qtext:
            return ''
        hits = []
        for nm in calc_names:
            if not nm:
                continue
            # 问题文本与分项名的公共关键词(2字以上)
            for k in (qtext[i:i + 2] for i in range(len(qtext) - 1)):
                if len(k) == 2 and k in nm and k not in ('影响', '分项', '计算', '图纸', '设计', '工程', '采用', '需要'):
                    if nm not in hits:
                        hits.append(nm)
                    break
            if len(hits) >= 3:
                break
        return ('（影响分项: ' + '、'.join(hits) + '）') if hits else ''

    issues = []
    for p in (problems or [])[:20]:
        qt = f"{p.get('类别', '')}: {p.get('问题', '')[:60]}"
        issues.append(f"□ {qt}{_link(qt)}"
                      f"{'（建议: ' + str(p.get('建议', ''))[:40] + '）' if p.get('建议') else ''}")
    for e in (estimated_items or [])[:20]:
        issues.append(f"□ 待核实(估算): {e.get('分项名称', '')} {e.get('工程量', '')}{e.get('单位', '')}"
                      f" — {str(e.get('计算式', ''))[:50]}")
    for p in (pending_items or [])[:20]:
        issues.append(f"□ 待提取(无证据): {p.get('分项名称', '')} — {str(p.get('计算式', p.get('原因', '')))[:50]}")
    for c in (pid.get('图纸问题候选', []) or [])[:20]:
        issues.append(f"□ {str(c)[:80]}{_link(str(c))}")
    if not issues:
        issues.append("无（本次识图未发现遗留问题）")
    sections['遗留问题'] = issues

    # 五、主要指标(造价自检)
    inds = pid.get('造价指标', {}) or {}
    if inds:
        ind_rows = [f"{k}: 实际 {v.get('实际', '')} (经验区间 {v.get('区间', '无')})"
                    f" 状态 {v.get('状态', '')}" for k, v in inds.items()]
    else:
        ind_rows = ["（算量质量报告生成，详见工程量计算书）"]
    sections['主要指标'] = ind_rows

    # 六、ABC 大项(审计反向: 前10大分项重点复核)
    abc = pid.get('ABC大项', []) or []
    if abc:
        abc_rows = [f"{i + 1}. {a.get('分项名称', '')} {a.get('工程量', '')}{a.get('单位', '')}"
                    f"（依据: {a.get('依据', '')[:40]}）" for i, a in enumerate(abc)]
    else:
        abc_rows = ["（算量质量报告生成）"]
    sections['ABC大项(重点复核)'] = abc_rows

    # 七、漏项检查(知识库工程类型常见项 + v6.9.5 自动对照结果)
    kc = pid.get('知识检查', {}) or {}
    checklist = kc.get('漏项检查', []) or []
    if checklist:
        sections['漏项检查(对照图纸核实)'] = checklist
    miss = pid.get('漏项对照', []) or []
    if miss:
        sections['漏项自动对照(疑似漏项)'] = [
            f"□ {mm} — 图纸{pid.get('工程性质', '')}类型常见项, 本次分项中未见对应, 请核实是否遗漏"
            for mm in miss]
    return sections


def export_note_xlsx(sections, path):
    """编制说明 → xlsx(条目式, 造价习惯排版)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = '编制说明'
    title_f = Font(name='微软雅黑', bold=True, size=16)
    sec_f = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
    sec_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    body_f = Font(name='微软雅黑', size=10)
    bd = Border(*[Side(style='thin')] * 4)
    ws.cell(row=1, column=1, value='编 制 说 明').font = title_f
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')
    r = 3
    for sec, lines in sections.items():
        ws.cell(row=r, column=1, value=sec).font = sec_f
        ws.cell(row=r, column=1).fill = sec_fill
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        for line in lines:
            ws.cell(row=r, column=1, value=line).font = body_f
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='top')
            r += 1
        r += 1
    ws.column_dimensions['A'].width = 110
    wb.save(path)
    return path


if __name__ == '__main__':
    # 自检: 用样例 pid 生成
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    demo = {'图纸元数据': {'图名': '样例'}, '专业类型': '房屋建筑与装饰工程',
            '工程性质': '新建', '面积区域': [{'面积_m2': 500, '面积来源': '闭合多段线'}],
            '构造层': [], '房间': [], '门窗': [], '钢结构': {}, '图纸问题候选': []}
    s = build_note(demo, [], [], [], [])
    p = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'output', '编制说明.xlsx')
    export_note_xlsx(s, p)
    print(f'自检输出: {p}')
