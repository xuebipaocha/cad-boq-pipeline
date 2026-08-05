"""
审图问题输出 - Excel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def export_review_excel(problems, output_path):
    """输出图纸问题清单到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = '图纸问题清单'
    
    # 样式
    hfill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    hfont = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    dfont = Font(name='微软雅黑', size=10)
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                 top=Side(style='thin'), bottom=Side(style='thin'))
    yfill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    
    ws.merge_cells('A1:G1')
    ws['A1'] = '图纸问题清单'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    headers = ['序号', '问题描述', '图纸位置', '类别', '影响造价', '严重程度', '建议']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = bdr
    
    for ri, p in enumerate(problems):
        r = ri + 4
        ws.cell(row=r, column=1, value=ri+1).font = dfont
        ws.cell(row=r, column=2, value=p.get('问题', '')).font = dfont
        ws.cell(row=r, column=3, value=p.get('位置', '')).font = dfont
        ws.cell(row=r, column=4, value=p.get('类别', '')).font = dfont
        ws.cell(row=r, column=5, value='是' if p.get('影响造价') else '否').font = dfont
        ws.cell(row=r, column=6, value=p.get('严重程度', '中')).font = dfont
        ws.cell(row=r, column=7, value=p.get('建议', '')).font = dfont
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = bdr
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
        if p.get('严重程度') == '高':
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = yfill
    
    for col, w in [('A',6),('B',40),('C',15),('D',12),('E',8),('F',8),('G',30)]:
        ws.column_dimensions[col].width = w
    
    wb.save(output_path)
    return output_path
