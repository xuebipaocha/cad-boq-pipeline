"""
填充国标清单模板（v5.16 重写）
- 表E.2.1 分部分项工程量清单计价表：序号/编码/名称/特征/单位/工程量/综合单价/合价
- 表E.2.2-3 综合单价分析表：每清单项一个块
    块结构（对照模板示例）：
      行1  表E.2.2-3 标题
      行2  工程名称
      行3  项目编码 | 项目名称 | 计量单位 | 工程量
      行4  项目特征描述
      行5  清单综合单价组成明细
      行6  表头(定额编号|定额项目名称|定额单位|数量)
      行7  单价/合价表头(人工费|材料费|机械费|管理费和利润 ×2)
      行8+ 定额明细行(每个定额子目一行, 单价+合价)
      小计行: 合价各列汇总
      未计价材料费行
      清单项目综合单价行: = 小计(人工+材料+机械+管理费利润) 即每清单单位的综合单价
    价格数据来自 step5 组价回填的 item['_price']（单一定额，含量=换算后定额量/清单量）。
"""
import sys, os, shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

def fill_boq_template(boq_items, template_path, output_path):
    """用清单数据填充国标模板"""
    shutil.copy2(template_path, output_path)

    if not os.path.exists(output_path):
        return False

    wb = load_workbook(output_path)
    dfont = Font(name='微软雅黑', size=10)
    bfont = Font(name='微软雅黑', bold=True, size=10)
    thin = Side(border_style="thin", color="000000")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: 表E.2.1 分部分项工程量清单计价表 ──
    if '表E.2.1 分部分项工程项目清单计价表' in wb.sheetnames:
        ws = wb['表E.2.1 分部分项工程项目清单计价表']

        # 清旧数据（R5以下）
        for row in range(5, ws.max_row + 1):
            for col in range(1, 11):
                try:
                    ws.cell(row=row, column=col).value = None
                except:
                    pass  # 合并单元格跳过

        # 填数据（按分部归类）— 含综合单价/合价(来自 _price)
        r = 5
        current_cat = None
        for item in boq_items:
            cat = item.get('section', item.get('category', '一般工程'))
            if current_cat != cat:
                if r > 5:
                    r += 1  # 分部间空行
                try:
                    ws.cell(row=r, column=3, value=cat).font = bfont
                except:
                    pass
                r += 1
                current_cat = cat

            p = item.get('_price') or {}
            unit_price = p.get('unit_price', '')
            total_price = p.get('total', '')
            for ci, v in [(1, item['seq']), (2, item.get('code', '')), (3, item['name']),
                          (4, item.get('features', '')), (6, item['unit']), (7, item['qty']),
                          (9, unit_price), (10, total_price)]:
                try:
                    ws.cell(row=r, column=ci, value=v).font = dfont
                except:
                    pass  # 合并单元格跳过
            r += 1

    # ── Sheet 2: 表E.2.2-3 综合单价分析表 ──
    if '表E.2.2-3 分部分项工程项目、技术措施清单综合单价分析表' in wb.sheetnames:
        ws2 = wb['表E.2.2-3 分部分项工程项目、技术措施清单综合单价分析表']

        # 清空数据 + 解除合并单元格(模板残留的合并会阻断写入)
        try:
            for mc in list(ws2.merged_cells.ranges):
                ws2.unmerge_cells(str(mc))
        except Exception:
            pass
        for row in range(1, ws2.max_row + 1):
            for col in range(1, 15):
                try:
                    ws2.cell(row=row, column=col).value = None
                except:
                    pass

        def _set(r, c, v, font=None, border=True):
            try:
                cell = ws2.cell(row=r, column=c, value=v)
                cell.font = font or dfont
                if border:
                    cell.border = bd
            except:
                pass

        # 每个清单项一个分析块（参照模板: 头13行 + 材料/机械明细可省）
        r = 1
        for item in boq_items:
            # 标题 + 工程名称
            _set(r, 1, "表E.2.2-3 分部分项工程项目、技术措施清单综合单价分析表", bfont); r += 1
            _set(r, 1, f"工程名称：{item.get('project', '')}", dfont); r += 1
            # 项目信息行
            _set(r, 1, "项目编码", bfont); _set(r, 3, item.get('code', ''))
            _set(r, 5, "项目名称", bfont); _set(r, 8, item['name'])
            _set(r, 10, "计量单位", bfont); _set(r, 11, item.get('unit', ''))
            _set(r, 13, "工程量", bfont); _set(r, 14, item.get('qty', ''))
            r += 1
            # 项目特征
            _set(r, 1, "项目特征描述", bfont); _set(r, 5, item.get('features', '')); r += 1
            # 明细标题
            _set(r, 1, "清单综合单价组成明细", bfont); r += 1
            # 表头
            _set(r, 1, "定额编号", bfont); _set(r, 2, "定额项目名称", bfont)
            _set(r, 3, "定额\n单位", bfont); _set(r, 4, "数量", bfont)
            _set(r, 5, "单价", bfont); _set(r, 10, "合价", bfont)
            r += 1
            _set(r, 5, "人工费", bfont); _set(r, 7, "材料费", bfont)
            _set(r, 8, "机械费", bfont); _set(r, 9, "管理费\n和利润", bfont)
            _set(r, 10, "人工费", bfont); _set(r, 11, "材料费", bfont)
            _set(r, 13, "机械费", bfont); _set(r, 14, "管理费\n和利润", bfont)
            r += 1

            # 定额明细行: 优先 sub_rows(多定额组合逐行), 否则单行(_price)
            p = item.get('_price') or {}
            qty = item.get('qty', 1) or 1
            sub_rows = p.get('sub_rows') or []
            if sub_rows:
                # 多定额组合: 逐行渲染子目
                for sr in sub_rows:
                    # v5.17: 借用/自补标记前缀(与真实工程一致: 借2-491 / 补001)
                    code_disp = sr.get('quota_code', '')
                    if sr.get('borrowed') and code_disp:
                        code_disp = f'借{code_disp}'
                    elif sr.get('supplement') and code_disp:
                        code_disp = f'补{code_disp}'
                    _set(r, 1, code_disp); _set(r, 2, sr.get('quota_name', ''))
                    _set(r, 3, sr.get('quota_unit', '')); _set(r, 4, sr.get('content', 1.0))
                    # 单价(每定额单位=合价/含量)
                    c = sr.get('content', 1.0) or 1.0
                    lab_u = sr.get('labor', 0) / c
                    mat_u = sr.get('material', 0) / c
                    mach_u = sr.get('machine', 0) / c
                    mp_u = sr.get('mgmt_profit', 0) / c
                    _set(r, 5, round(lab_u, 2)); _set(r, 7, round(mat_u, 2))
                    _set(r, 8, round(mach_u, 2)); _set(r, 9, round(mp_u, 2))
                    # 合价(每清单单位)
                    _set(r, 10, sr.get('labor', 0)); _set(r, 11, sr.get('material', 0))
                    _set(r, 13, sr.get('machine', 0)); _set(r, 14, round(sr.get('mgmt_profit', 0), 2))
                    r += 1
                # 小计 = 各子目合价汇总
                _set(r, 1, "小计", bfont)
                _set(r, 10, round(sum(s.get('labor', 0) for s in sub_rows), 2))
                _set(r, 11, round(sum(s.get('material', 0) for s in sub_rows), 2))
                _set(r, 13, round(sum(s.get('machine', 0) for s in sub_rows), 2))
                _set(r, 14, round(sum(s.get('mgmt_profit', 0) for s in sub_rows), 2))
                r += 1
                # 未计价材料费
                _set(r, 1, "未计价材料费", bfont); _set(r, 10, 0); r += 1
                # 清单项目综合单价 = 小计合计(每清单单位)
                comp = p.get('unit_price', 0) or 0
                _set(r, 1, "清单项目综合单价", bfont); _set(r, 10, comp, bfont)
                r += 1
            elif p:
                quota_qty = p.get('quota_qty', qty)
                content = round(quota_qty / qty, 4) if qty else 1.0
                lab, mat, mach = p.get('labor', 0), p.get('material', 0), p.get('machine', 0)
                mgmt, prof = p.get('mgmt', 0), p.get('profit', 0)
                mgmt_profit = mgmt + prof
                _set(r, 1, p.get('quota_code', '')); _set(r, 2, p.get('quota_name', ''))
                _set(r, 3, p.get('quota_unit', '')); _set(r, 4, content)
                # 单价(每定额单位)
                _set(r, 5, lab); _set(r, 7, mat); _set(r, 8, mach); _set(r, 9, round(mgmt_profit, 2))
                # 合价(每清单单位 = 单价×含量)
                _set(r, 10, round(lab * content, 2)); _set(r, 11, round(mat * content, 2))
                _set(r, 13, round(mach * content, 2)); _set(r, 14, round(mgmt_profit * content, 2))
                r += 1
                # 小计
                _set(r, 1, "小计", bfont)
                _set(r, 10, round(lab * content, 2)); _set(r, 11, round(mat * content, 2))
                _set(r, 13, round(mach * content, 2)); _set(r, 14, round(mgmt_profit * content, 2))
                r += 1
                # 未计价材料费
                _set(r, 1, "未计价材料费", bfont); _set(r, 10, 0); r += 1
                # 清单项目综合单价 = 小计合计(每清单单位)
                comp = round(lab * content + mat * content + mach * content + mgmt_profit * content, 2)
                _set(r, 1, "清单项目综合单价", bfont); _set(r, 10, comp, bfont)
                r += 1
            else:
                # 无价格: 从 mapped_quotas 填候选(编号/名称/单位)
                for q in (item.get('mapped_quotas') or [])[:5]:
                    _set(r, 1, q.get('quota_code', '')); _set(r, 2, q.get('item_name', ''))
                    _set(r, 3, q.get('unit', '')); _set(r, 4, 1)
                    r += 1
                _set(r, 1, "小计", bfont); r += 1
                _set(r, 1, "未计价材料费", bfont); r += 1
                _set(r, 1, "清单项目综合单价", bfont); r += 1

            # 块间空行
            r += 1

    wb.save(output_path)
    return True
