"""Step 5: 组价 — 准确性增强版

新增能力：
- 优先使用 Step4 的清单-定额映射候选。
- 定额匹配带置信度和复核提示。
- 清单单位 -> 定额单位自动换算。
- 管理费、利润、措施费各自按自身 base_calc 判断取费基数。
- 费率专业名通过映射表查找，修复市政/安装/园林费率为0的问题。
"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(BASE_DIR)
sys.path.insert(0, BASE_DIR)
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


def _rate(info):
    return (info.get('rate') or 0) / 100


# v6.9.1: 定额成本基数校准规则缓存(见 quota_calibrate.py)
_CALIBRATION = None


def _load_calibration():
    global _CALIBRATION
    if _CALIBRATION is None:
        try:
            p = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                             'data', 'quota_calibration.json')
            with open(p, encoding='utf-8') as f:
                _CALIBRATION = json.load(f).get('规则', [])
        except Exception:
            _CALIBRATION = []
    return _CALIBRATION


def _base_factor(info, db):
    return db.base_factor_from_calc(info.get('base_calc', ''))


def _load_combos():
    """加载多定额组合规则。"""
    import json as _json
    p = os.path.join(SKILL_DIR, 'data', 'quota_combos.json')
    try:
        with open(p, encoding='utf-8') as f:
            return _json.load(f)
    except Exception:
        return {}


# v5.17: 自补定额序号(补001, 补002 ...)
_SUPPLEMENT_SEQ = [1]


def _find_combo_rule(name, combos):
    """按清单分项名匹配组合规则, 返回规则dict或None。"""
    for rule in combos.values():
        if not isinstance(rule, dict):
            continue
        for kw in rule.get('匹配', []):
            if kw in name:
                return rule
    return None


def _match_conditions(sub, features_text):
    """子目条件判定: 无条件=固定组合; 有条件=特征文本须含任一关键词。"""
    conds = sub.get('条件') or []
    if not conds:
        return True
    if not features_text:
        return False
    return any(c in features_text for c in conds)


# v6.9.3: 材质词表 — 组价按材质匹配定额(统一逻辑: 清单特征材质 → 定额材料/工作内容)
# 用户确认: 所有项目一个逻辑, 不要项目级特判
MATERIAL_KWS = ['钢质', '塑钢', '木质', '铝合金', '断桥铝', '防火', '推拉', '平开',
                '铸铁', 'PPR', 'UPVC', '镀锌钢管', '无缝钢管', '衬塑', '不锈钢',
                '乳胶漆', '真石漆', '氟碳漆', '瓷砖', '石材', '木地板', '卷材', '聚氨酯',
                '加气块', '实心砖', '多孔砖', '陶粒混凝土']


def _extract_material_kw(name, features_text=''):
    """从清单名+特征提取材质词(第一个命中)。"""
    hay = f'{name} {features_text or ""}'
    for w in MATERIAL_KWS:
        if w in hay:
            return w
    return None


def _find_quota_by_material(kw, specialty, db, name):
    """材质词+类型词硬约束检索定额(名称含材质词; 门/窗类结合类型词)。
    统一逻辑: 清单特征材质 → 定额材料/工作内容(用户确认, 土建+安装通用)。"""
    try:
        from pipeline.db import get_liaoning_conn
        # v6.9.3: 类型词结合 — '塑钢窗更换' 应匹配 8-71 塑钢成品窗而非 8-10 塑钢门
        type_kw = ''
        if '窗' in name and '门' not in name:
            type_kw = '窗'
        elif '门' in name and '窗' not in name:
            type_kw = '门'
        conn = get_liaoning_conn()
        try:
            cat = db.SPECIALTY_CATEGORY.get(specialty, [specialty])
            cat_cond = ' OR '.join(['category LIKE ? OR sub_category LIKE ?' for _ in cat])
            params = []
            for c in cat:
                params.extend([f'%{c}%', f'%{c}%'])
            if type_kw:
                rows = conn.execute(
                    f"SELECT * FROM quota_items WHERE item_name LIKE ? AND item_name LIKE ? "
                    f"AND ({cat_cond}) AND base_price>0 LIMIT 15",
                    (f'%{kw}%', f'%{type_kw}%') + tuple(params)).fetchall()
            else:
                rows = conn.execute(
                    f"SELECT * FROM quota_items WHERE item_name LIKE ? AND ({cat_cond}) AND base_price>0 LIMIT 15",
                    (f'%{kw}%',) + tuple(params)).fetchall()
            conn.close()
        except Exception:
            conn.close()
            rows = []
        if rows:
            cands = sorted(rows, key=lambda x: (len(x['item_name'] or ''), -(x['base_price'] or 0)))
            q = dict(cands[0])
            q['unit'] = db.infer_unit(q.get('item_name'), q.get('unit'))
            q['_score'] = 0.8
            q['_confidence'] = '高'
            q['_match_method'] = f'material:{kw}{type_kw}'
            return q
    except Exception:
        pass
    return None


def _pick_quotas(item, specialty, db, combos, features_text=''):
    """多定额组合: 返回 [{quota, content, note}...]。无规则时回退单定额。
    features_text: 清单特征/做法表/施工说明合并文本, 用于子目条件判定。
    """
    name = item.get('source_name') or item.get('name') or ''
    mapped = item.get('mapped_quotas') or []
    # v6.9.3: 材质匹配优先(统一逻辑) — 只从【分项自身】名称+特征提取材质词
    # (item['features'] 是该分项特征, features_text 是全图文本 — 全图提取会让
    # 任意项错配到图内其他材质定额, 如'门窗更换(汇总)'错配钢质防盗门8-14)
    _mkw = _extract_material_kw(name, item.get('features') or '')
    mq = _find_quota_by_material(_mkw or '', specialty, db, name) if _mkw else None
    if mq:
        return [{'quota': mq, 'content': 1.0, 'note': f'材质匹配定额({mq.get("_match_method", "")})'}]
    # v6.9.2/3: 门窗泛化项(更换/拆除/洞口面积, 含'未注明材质'类)不走组合规则 —
    # 组合规则'拆除门窗'/'拆除通用'子目检索会漂移到高价装饰定额(17-229 门窗框装饰
    # 线条 5630元/m²、16-2 实心墙 3806元/m² 实测); 无材质词命中的门窗项统一自补
    if ('门窗' in name or ('门' in name or '窗' in name)) \
            and any(k in name for k in ('更换', '拆除', '洞口面积')):
        rule = None
        _SUPPLEMENT_SEQ[0] += 1
        _sq = {
            'quota_code': f'{_SUPPLEMENT_SEQ[0]:03d}',
            'item_name': f'{name}（自补定额）',
            'unit': item.get('unit') or '',
            'base_price': 0, 'labor_cost': 0, 'material_cost': 0, 'machine_cost': 0,
            '_score': 0, '_confidence': '待确认', '_match_method': 'supplement',
            '_supplement': True,
        }
        return [{'quota': _sq, 'content': 1.0, 'note': '自补定额，人材机待补充（门窗泛化项防错配）'}]
    rule = _find_combo_rule(name, combos)
    if rule:
        subs = []
        for sub in rule.get('子目', []):
            if not _match_conditions(sub, features_text):
                continue  # 特征不满足条件 → 不套该子目
            kw = sub.get('关键词', '')
            # 子目检索: 优先在映射候选里找, 否则按名称模糊检索
            q = None
            for m in mapped:
                if kw in (m.get('item_name') or ''):
                    q = m
                    break
            if not q:
                try:
                    # 硬约束检索: 名称必须包含关键词(避免相似度漂移)
                    from pipeline.db import get_liaoning_conn
                    conn = get_liaoning_conn()
                    try:
                        cat = db.SPECIALTY_CATEGORY.get(specialty, [specialty])
                        cat_cond = ' OR '.join(['category LIKE ? OR sub_category LIKE ?' for _ in cat])
                        params = []
                        for c in cat:
                            params.extend([f'%{c}%', f'%{c}%'])
                        rows = conn.execute(
                            f"SELECT * FROM quota_items WHERE item_name LIKE ? AND ({cat_cond}) AND base_price>0 LIMIT 20",
                            (f'%{kw}%',) + tuple(params)).fetchall()
                        borrow = False
                    except Exception:
                        rows = []
                        borrow = False
                    if not rows:
                        # v5.17: 跨册借用 — 本专业册无合适子目时放宽到全库检索(标记"借")
                        rows = conn.execute(
                            "SELECT * FROM quota_items WHERE item_name LIKE ? AND base_price>0 LIMIT 20",
                            (f'%{kw}%',)).fetchall()
                        borrow = bool(rows)
                    conn.close()
                    if rows:
                        # 取基价最高的前几个, 按名称最短优先(更接近主条目)
                        cands = sorted(rows, key=lambda x: (len(x['item_name'] or ''), -(x['base_price'] or 0)))
                        q = dict(cands[0])
                        q['unit'] = db.infer_unit(q.get('item_name'), q.get('unit'))
                        q['_score'] = 0.9
                        q['_confidence'] = '高'
                        q['_match_method'] = 'combo_hard'
                        if borrow:
                            q['_borrowed'] = True  # 跨册借用标记
                    else:
                        quotas = db.find_quota(kw, category=specialty, top_n=3)
                        q = quotas[0] if quotas else None
                except Exception:
                    q = None
            if q:
                q = dict(q)
                q['_selected_from'] = q.get('_match_method', 'combo')
                subs.append({'quota': q, 'content': float(sub.get('含量', 1.0)),
                             'note': sub.get('备注', '')})
            else:
                # v5.17: 自补定额 — 本册/跨册均无合适子目, 生成自补条目(人材机待人工补充)
                q = {
                    'quota_code': f'{_SUPPLEMENT_SEQ[0]:03d}',
                    'item_name': f'{name}（自补定额）',
                    'unit': item.get('unit') or '',
                    'base_price': 0, 'labor_cost': 0, 'material_cost': 0, 'machine_cost': 0,
                    '_score': 0, '_confidence': '待确认', '_match_method': 'supplement',
                    '_supplement': True,
                }
                _SUPPLEMENT_SEQ[0] += 1
                subs.append({'quota': q, 'content': float(sub.get('含量', 1.0)),
                             'note': '自补定额，人材机待补充'})
        if subs:
            return subs
    # 回退: 单定额(原逻辑)
    q = None
    # v6.9.2: 拆除类换安装定额 — 真实工程'换'惯例(拆除=安装×人材机系数 0.5/0/0),
    # 直接检索拆除定额常无结果, 按规则换对应安装定额(显示'换'标记)
    if '拆除' in name and not mapped:
        try:
            from pipeline.db import get_liaoning_conn
            _conn = get_liaoning_conn()
            _base_kw = name.replace('拆除', '安装')
            _rows = _conn.execute(
                "SELECT * FROM quota_items WHERE item_name LIKE ? AND base_price>0 LIMIT 10",
                (f'%{_base_kw}%',)).fetchall()
            _conn.close()
            if _rows:
                _c = dict(sorted(_rows, key=lambda x: (len(x['item_name'] or ''), -(x['base_price'] or 0)))[0])
                _c['unit'] = db.infer_unit(_c.get('item_name'), _c.get('unit'))
                _c['_score'] = 0.6
                _c['_confidence'] = '中'
                _c['_match_method'] = 'exchange'
                _c['_exchange'] = {'基准': _base_kw, '系数': {'人工': 0.5, '材料': 0.0, '机械': 0.0},
                                   '说明': '拆除=安装×人工0.5材料0机械0(行业惯例, ⚠️待核辽宁定额)'}
                q = _c
        except Exception:
            q = None
    # v6.9.1: 泛化项(门窗更换/拆除/洞口面积)强制自补 — 模糊匹配实测错配到
    # 高价装饰定额('门窗框小型构件装饰线条增加费' 5630元/m²), 自补最诚实
    if '门窗' in name and any(k in name for k in ('更换', '拆除', '洞口面积')):
        q = {
            'quota_code': f'{_SUPPLEMENT_SEQ[0]:03d}',
            'item_name': f'{name}（自补定额）',
            'unit': item.get('unit') or '',
            'base_price': 0, 'labor_cost': 0, 'material_cost': 0, 'machine_cost': 0,
            '_score': 0, '_confidence': '待确认', '_match_method': 'supplement',
            '_supplement': True,
        }
        _SUPPLEMENT_SEQ[0] += 1
        return [{'quota': q, 'content': 1.0, 'note': '自补定额，人材机待补充（泛化项防错配）'}]
    if mapped:
        q = mapped[0]
        q['_selected_from'] = 'mapping'
    else:
        unit = item.get('unit') or item.get('list_unit') or ''
        quotas = db.find_quota(name, category=specialty, top_n=5,
                               list_code=(item.get('code') if item.get('code') != '（待匹配）' else None),
                               unit=unit)
        q = quotas[0] if quotas else {}
        if q:
            q['_selected_from'] = q.get('_match_method', 'similarity')
    if not q:
        # v5.17: 自补定额 — 无任何匹配时生成自补条目
        q = {
            'quota_code': f'{_SUPPLEMENT_SEQ[0]:03d}',
            'item_name': f'{name}（自补定额）',
            'unit': item.get('unit') or '',
            'base_price': 0, 'labor_cost': 0, 'material_cost': 0, 'machine_cost': 0,
            '_score': 0, '_confidence': '待确认', '_match_method': 'supplement',
            '_supplement': True,
        }
        _SUPPLEMENT_SEQ[0] += 1
        return [{'quota': q, 'content': 1.0, 'note': '自补定额，人材机待补充'}]
    return [{'quota': q, 'content': 1.0, 'note': ''}]


def _pick_quota(item, specialty, db):
    """单定额回退(兼容旧调用)。"""
    subs = _pick_quotas(item, specialty, db, {})
    return subs[0]['quota'] if subs else {}


# v5.19: 主材价格匹配 — 真实价优先(大连信息价/真实询价), 无则经验价
_MATERIAL_NAMES_CACHE = None


def _load_material_names(db):
    """加载价目表全部材料名(缓存, 按长度降序)。"""
    global _MATERIAL_NAMES_CACHE
    if _MATERIAL_NAMES_CACHE is None:
        try:
            from pipeline.db import get_liaoning_conn
            conn = get_liaoning_conn()
            names = [r[0] for r in conn.execute(
                "SELECT material_name FROM material_prices ORDER BY LENGTH(material_name) DESC")]
            conn.close()
            _MATERIAL_NAMES_CACHE = [n for n in names if n and len(n) >= 2]
        except Exception:
            _MATERIAL_NAMES_CACHE = []
    return _MATERIAL_NAMES_CACHE


def _match_main_material(item, features_text, db):
    """从清单名+该清单项特征中匹配主材, 返回 {name, unit, price, source, spec} 或 None。
    v6.3 B3: ①规格优先(特征/设计说明含 800×800/DN25/C30 → 匹配"材料名 规格")
             ②双向主干匹配(材料名主干↔清单名, 如 '地砖 300×300' ↔ '地砖地面')。
    """
    import re as _re
    name_text = f"{item.get('source_name') or ''} {item.get('name') or ''}"
    feat_text = item.get('features') or ''
    combined = name_text + ' ' + feat_text

    # ① 规格优先
    spec_pats = [r'(\d{2,4})\s*[×xX*]\s*(\d{2,4})', r'(DN\d{2,3})', r'\b(C\d{2})\b', r'\b(M\d+(?:\.\d)?)\b']
    specs = []
    for pat in spec_pats:
        for m in _re.finditer(pat, combined):
            s = m.group(0)
            if s not in specs:
                specs.append(s)
    if specs:
        for mname in _load_material_names(db):
            if mname == '水':
                continue
            # 材料名含规格(如 '地砖 800×800') 且 材料主干在清单文本中
            if any(s in mname for s in specs):
                trunk = mname.split(' ')[0].strip()
                if trunk and trunk in combined:
                    try:
                        rows = db.find_material_price(mname, top_n=1)
                        if rows:
                            return rows[0]
                    except Exception:
                        continue

    # ② 双向主干匹配: 材料名在清单文本 或 材料主干(空格前)在清单文本
    for mname in _load_material_names(db):
        if mname == '水':
            continue
        if mname in combined:
            try:
                rows = db.find_material_price(mname, top_n=1)
                if rows:
                    return rows[0]
            except Exception:
                continue
    for mname in _load_material_names(db):
        if mname == '水':
            continue
        trunk = mname.split(' ')[0].strip()
        if len(trunk) >= 2 and trunk in combined:
            try:
                rows = db.find_material_price(mname, top_n=1)
                if rows:
                    return rows[0]
            except Exception:
                continue

    # ③ v6.9.3: 材质词模糊匹配 — '塑钢、推拉窗更换' 含'塑钢'但子串不连续,
    # 按材质词 LIKE 匹配 material_prices('塑钢推拉窗' 239元/m²); 类型词(门/窗)优先
    _MATERIAL_WORDS = ['塑钢', '钢质', '木质', '铝合金', '断桥铝', '铸铁', '不锈钢']
    _hit_word = next((w for w in _MATERIAL_WORDS if w in combined), None)
    if _hit_word:
        _type = '窗' if ('窗' in combined and '门' not in combined) else ('门' if '门' in combined else '')
        _cands = [m for m in _load_material_names(db) if _hit_word in m]
        if _type:
            _cands = [m for m in _cands if _type in m] or _cands
        if _cands:
            _cands.sort(key=len)
            try:
                rows = db.find_material_price(_cands[0], top_n=1)
                if rows:
                    return rows[0]
            except Exception:
                pass
    return None


def _material_price_of(item, features_text, db, qty=1.0):
    """主材计价: 返回 {main_material, main_price, main_unit, main_source, dim_ok} —
    main_price = 单价(每主材单位)。dim_ok = 主材单位与清单单位维度是否一致:
    一致(如 m3→m3)才计入综合单价; 不一致(如 t→m3)标注待人工补含量, 不计入。
    真实价优先(find_material_price 已含 dalian/real_boq/experience 全部来源)。
    """
    m = _match_main_material(item, features_text, db)
    if not m:
        return {'main_material': '', 'main_price': 0.0, 'main_unit': '', 'main_source': '', 'dim_ok': False}
    # 单位维度判断
    dim_ok = False
    try:
        list_unit = item.get('unit') or ''
        d1 = db.unit_dimension(m.get('unit', ''))
        d2 = db.unit_dimension(list_unit)
        dim_ok = bool(d1 and d2 and d1 == d2)
    except Exception:
        dim_ok = False
    return {'main_material': m['material_name'], 'main_price': float(m['price']),
            'main_unit': m.get('unit', ''), 'main_source': m.get('source', ''),
            'dim_ok': dim_ok}


def run(boq_json, output_dir):
    print('='*50)
    print('Step 5: 组价 — 增强取费')
    print('='*50)
    from pipeline import db

    with open(boq_json, 'r', encoding='utf-8') as f:
        boq_items = json.load(f)

    # v5.17: 特征文本(构造层+施工说明+表格) → 子目条件判定依据
    features_text = ''
    recog_path = os.path.join(output_dir, '识图结果.json')
    if os.path.exists(recog_path):
        try:
            with open(recog_path, encoding='utf-8') as f:
                recog = json.load(f)
            parts = []
            for layer in recog.get('构造层') or []:
                parts.append(' '.join(str(layer.get(k) or '') for k in ('名称', '材料', '部位', '厚度来源')))
            for note in recog.get('施工说明') or []:
                parts.append(str(note))
            for tbl in recog.get('表格') or []:
                parts.append(str(tbl))
            features_text = ' '.join(parts)
        except Exception:
            pass

    specialty = '市政工程'
    recog_path = os.path.join(output_dir, '识图结果.json')
    if os.path.exists(recog_path):
        with open(recog_path, encoding='utf-8') as f:
            specialty = json.load(f).get('专业类型', '市政工程')
    print(f'  专业: {specialty}')

    fees = db.fee_rates_for_specialty(specialty)
    env_info = fees['文明施工和环境保护费']
    rainy_info = fees['雨季施工费']
    winter_info = fees['冬季施工费']
    mgmt_info = fees['企业管理费']
    profit_info = fees['利润']
    safety_rate = _rate(fees['安全施工费'])
    reg_rate = _rate(fees['规费'])
    vat_rate = _rate(fees['增值税'])

    print(f"  管理费费率: {_rate(mgmt_info)*100:.2f}% | 来源:{mgmt_info.get('profession','')}")
    print(f"  利润率: {_rate(profit_info)*100:.2f}% | 来源:{profit_info.get('profession','')}")

    wb = Workbook()
    ws = wb.active
    ws.title = '已组价清单'

    hf = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hft = Font(name='微软雅黑', bold=True, size=9, color='FFFFFF')
    df = Font(name='微软雅黑', size=9)
    bd = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    ca = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['序号','项目编码','项目名称','清单单位','清单量','定额编号','定额名称','定额单位','换算后定额量','匹配分','置信度','人工费','材料费','机械费','管理费','利润','综合单价','合价','备注']
    ws.cell(row=1,column=1,value='已组价工程量清单（增强准确性）').font = Font(name='微软雅黑', bold=True, size=14)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(headers))
    ws.cell(row=2,column=1,value=f"专业:{specialty} | 管理费:{_rate(mgmt_info)*100:.2f}% | 利润:{_rate(profit_info)*100:.2f}% | 安全施工:{safety_rate*100:.2f}% | 规费:{reg_rate*100:.2f}% | 增值税:{vat_rate*100:.2f}%").font = Font(name='微软雅黑', size=8, color='666666')
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(headers))

    for i,h in enumerate(headers,1):
        c=ws.cell(row=4,column=i,value=h); c.font=hft; c.fill=hf; c.alignment=ca; c.border=bd

    total_labor = total_material = total_machine = total_mgmt = total_profit = total_sum = 0
    low_conf = 0
    r = 5

    for item in boq_items:
        name = item.get('name','')
        source_name = item.get('source_name') or name
        qty = item.get('qty',0) or 0
        list_unit = db.infer_unit(source_name, item.get('unit') or item.get('list_unit') or '')

        # v5.16/v5.17: 多定额组合(特征驱动: 子目条件按图纸特征判定)
        combos = _load_combos()
        item_ft = item.get('features_text') or features_text  # 优先 step4 同源特征
        subs = _pick_quotas(item, specialty, db, combos, features_text=item_ft)
        # 子目明细(供分析表)
        sub_rows = []
        # 汇总各子目
        agg = {'lab': 0.0, 'mat': 0.0, 'mach': 0.0, 'mgmt': 0.0, 'prof': 0.0,
               'qty': 0.0, 'score': 0.0, 'conf': '', 'qc': '', 'qname': '',
               'quota_unit': '', 'conv_note': '', 'base': 0.0}
        if not subs:
            subs = [{'quota': {}, 'content': 1.0, 'note': '未匹配定额'}]
        multi = len(subs) > 1
        conv_factor = 1.0  # 首个子目换算因子(供回填)
        for si, sub in enumerate(subs):
            q = sub.get('quota') or {}
            content = float(sub.get('content', 1.0))
            qc = q.get('quota_code','')
            qname = q.get('item_name','')
            quota_unit = db.infer_unit(qname, q.get('unit',''))
            quota_qty, conv_factor, conv_note = db.quota_qty_from_list_qty(qty, list_unit, quota_unit)
            # 含量 = 换算后定额量 × 子目系数
            sub_qty = quota_qty * content
            score = q.get('_score', 0) or 0
            confidence = q.get('_confidence', '待确认') or '待确认'
            if score < 0.45:
                low_conf += 1

            lab = (q.get('labor_cost') or 0) * content
            mat = (q.get('material_cost') or 0) * content
            mach = (q.get('machine_cost') or 0) * content
            base = q.get('base_price', 0) or 0
            if lab == 0 and mat == 0 and mach == 0 and base > 0:
                lab, mat, mach = db.decompose_cost(base, q.get('category') or specialty)
                lab, mat, mach = lab * content, mat * content, mach * content

            # v6.9.1: 定额成本基数校准 — 库重建丢失计量系数(成本按10/100倍单位编制),
            # 规则表按(名称关键词,单位)查基数, 人材机除以基数
            # (实测根因: 抹灰 12-1 成本按 100m² 编制, 未校准前人工 1801.58 元/m²)
            try:
                _cal = _load_calibration()
                _base = None
                for _rule in _cal:
                    _kws = _rule.get('关键词组') or []
                    if any(_kw in (qname or '') for _kw in _kws) and quota_unit == _rule.get('单位'):
                        _base = _rule['基数']
                        break
                if _base and _base != 1:
                    lab = lab / _base
                    mat = mat / _base
                    mach = mach / _base
                    base = base / _base
            except Exception:
                pass

            # v6.9.2: 换定额系数应用(拆除=安装×人工0.5材料0机械0)
            _ex = (q or {}).get('_exchange')
            if _ex and _ex.get('系数'):
                _k = _ex['系数']
                lab = lab * float(_k.get('人工', 1.0) or 1.0)
                mat = mat * float(_k.get('材料', 1.0) or 1.0)
                mach = mach * float(_k.get('机械', 1.0) or 1.0)

            mgmt_base = (lab + mach) * _base_factor(mgmt_info, db)
            profit_base = (lab + mach) * _base_factor(profit_info, db)
            mgmt = mgmt_base * _rate(mgmt_info)
            prof = profit_base * _rate(profit_info)
            unit_cost = lab + mat + mach + mgmt + prof  # 每清单单位成本(子目部分)
            sub_total = unit_cost * qty

            agg['lab'] += lab; agg['mat'] += mat; agg['mach'] += mach
            agg['mgmt'] += mgmt; agg['prof'] += prof
            agg['qty'] += sub_qty
            agg['score'] = max(agg['score'], score)
            agg['conf'] = confidence if si == 0 else agg['conf']
            agg['qc'] = qc if si == 0 else agg['qc']
            agg['qname'] = qname if si == 0 else agg['qname']
            agg['quota_unit'] = quota_unit if si == 0 else agg['quota_unit']
            agg['conv_note'] = conv_note if si == 0 else agg['conv_note']
            agg['base'] += base * content

            sub_rows.append({
                'quota_code': qc, 'quota_name': qname, 'quota_unit': quota_unit,
                'content': round(content, 4), 'sub_qty': round(sub_qty, 4),
                'labor': round(lab, 2), 'material': round(mat, 2), 'machine': round(mach, 2),
                'mgmt_profit': round(mgmt + prof, 2),
                'note': sub.get('note', ''),
                'borrowed': bool(q.get('_borrowed')),
                'supplement': bool(q.get('_supplement')),
                'exchange': bool(q.get('_exchange')),
            })

        lab, mat, mach = agg['lab'], agg['mat'], agg['mach']
        mgmt, prof = agg['mgmt'], agg['prof']
        quota_qty = agg['qty']
        unit_price = agg['lab'] + agg['mat'] + agg['mach'] + agg['mgmt'] + agg['prof']
        total = unit_price * qty
        qc, qname, quota_unit = agg['qc'], agg['qname'], agg['quota_unit']
        conv_note = agg['conv_note']
        base = agg['base']
        score, confidence = agg['score'], agg['conf']

        # v5.19: 主材价格计入(真实价优先, 无则经验价) — 单位维度一致才计入综合单价
        mp = _material_price_of(item, item_ft, db, qty)
        main_mat = mp['main_material']
        main_price = mp['main_price']
        if main_price > 0 and mp['dim_ok']:
            mat += main_price
            unit_price += main_price
            total += main_price * qty

        total_labor += lab * qty
        total_material += mat * qty
        total_machine += mach * qty
        total_mgmt += mgmt * qty
        total_profit += prof * qty
        total_sum += total

        notes = []
        if item.get('is_substitute'): notes.append('清单待复核')
        if confidence in ('低','待确认'): notes.append('定额低置信度')
        if '未换算' in conv_note or '不一致' in conv_note: notes.append(conv_note)
        if not qc: notes.append('未匹配定额')
        if base and (lab == 0 or mat == 0 or mach == 0): notes.append('人材机可能不完整')
        if multi: notes.append(f'组合{len(subs)}项定额')
        # v5.17: 借用/自补标记
        if any(s.get('borrowed') for s in sub_rows): notes.append('含跨册借用定额')
        if any(s.get('supplement') for s in sub_rows): notes.append('含自补定额')
        if any(s.get('exchange') for s in sub_rows): notes.append('换定额(拆除=安装×0.5/0/0, ⚠️行业惯例待核)')
        # v5.19: 主材标记
        if main_mat:
            src_txt = {'dalian': '大连信息价', 'real_boq': '真实询价', 'experience': '经验价'}.get(mp['main_source'], mp['main_source'])
            if mp['dim_ok']:
                notes.append(f'主材[{main_mat[:20]} {main_price:.2f}/{mp["main_unit"] or "-"}({src_txt})]')
            else:
                notes.append(f'主材[{main_mat[:20]} {main_price:.2f}/{mp["main_unit"] or "-"}({src_txt})]待补含量')

        # v6.9: 单价合理性校验(价格直觉) — 综合单价超出经验区间±30% → 警告
        try:
            import os as _os, json as _json
            _kp = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
                                'data', 'knowledge', 'unit_prices.json')
            if _os.path.exists(_kp):
                with open(_kp, encoding='utf-8') as _f:
                    _prices = _json.load(_f)
                for _k, _e in _prices.items():
                    if _k.startswith('_'):
                        continue
                    if any(kw in (name or '') for kw in _e.get('关键词', [])):
                        lo, hi = _e['区间']
                        if unit_price > 0 and not (lo * 0.7 <= unit_price <= hi * 1.3):
                            notes.append(f'⚠单价{unit_price:.0f}{_e["单位"]}超出经验区间'
                                         f'({lo}-{hi}, {_e.get("年份", "")}口径, 来源{_e.get("来源", "")[:18]})')
                        break
        except Exception:
            pass

        # v5.17: 定额编号显示前缀(借/补)
        qc_disp = qc
        if qc:
            if any(s.get('borrowed') for s in sub_rows):
                qc_disp = f'借{qc}'
            elif any(s.get('supplement') for s in sub_rows):
                qc_disp = f'补{qc}'

        vals = [item.get('seq',''), item.get('code',''), name, list_unit, qty, qc_disp, qname, quota_unit, round(quota_qty,4), round(score,4), confidence,
                round(lab,2), round(mat,2), round(mach,2), round(mgmt,2), round(prof,2), round(unit_price,2), round(total,2), '；'.join(notes)]
        for ci,v in enumerate(vals,1):
            cell=ws.cell(row=r,column=ci,value=v); cell.font=df; cell.border=bd; cell.alignment=ca if ci not in (3,7,19) else Alignment(vertical='center', wrap_text=True)
        r += 1

        # v5.16: 组价结果回填 item → 供综合单价分析表使用
        item['_price'] = {
            'quota_code': qc, 'quota_name': qname, 'quota_unit': quota_unit,
            'quota_qty': quota_qty, 'conv_factor': conv_factor, 'conv_note': conv_note,
            'qty': qty, 'list_unit': list_unit,
            'labor': round(lab, 2), 'material': round(mat, 2), 'machine': round(mach, 2),
            'mgmt': round(mgmt, 2), 'profit': round(prof, 2),
            'unit_price': round(unit_price, 2), 'total': round(total, 2),
            'score': score, 'confidence': confidence, 'notes': notes,
            'base_price': base or 0,
            'multi': multi,
            'sub_rows': sub_rows,  # 多定额子目明细(分析表逐行)
            # v5.19: 主材(未计价材料)
            'main_material': main_mat,
            'main_price': round(main_price, 2),
            'main_unit': mp['main_unit'],
            'main_source': mp['main_source'],
        }

    r += 1
    ws.cell(row=r,column=1,value='分部分项合计').font=Font(name='微软雅黑', bold=True, size=10)
    for c in range(1,len(headers)+1): ws.cell(row=r,column=c).border=bd
    for ci,v in [(12,round(total_labor,2)),(13,round(total_material,2)),(14,round(total_machine,2)),(15,round(total_mgmt,2)),(16,round(total_profit,2)),(18,round(total_sum,2))]:
        ws.cell(row=r,column=ci,value=v).font=Font(name='微软雅黑', bold=True, size=9); ws.cell(row=r,column=ci).alignment=ca

    base_mei_env = (total_labor + total_machine) * _base_factor(env_info, db)
    base_mei_rainy = (total_labor + total_machine) * _base_factor(rainy_info, db)
    base_mei_winter = (total_labor + total_machine) * _base_factor(winter_info, db)
    # v4.0: 安全施工费基数按 fee_rates 表 base_calc 口径（税前分部分项费）, 不再无条件用总额
    safety_base_calc = fees['安全施工费'].get('base_calc', '')
    if '人工' in safety_base_calc or '机械' in safety_base_calc:
        safety_base = (total_labor + total_machine) * _base_factor(fees['安全施工费'], db)
    else:
        safety_base = total_sum
    safety_fee = safety_base * safety_rate
    env_fee = base_mei_env * _rate(env_info)
    rainy_fee = base_mei_rainy * _rate(rainy_info)
    winter_fee = base_mei_winter * _rate(winter_info)
    total_measures = safety_fee + env_fee + rainy_fee + winter_fee

    for label, val in [('安全施工费', safety_fee), ('文明施工和环境保护费', env_fee), ('雨季施工费', rainy_fee), ('冬季施工费', winter_fee)]:
        r += 1
        ws.cell(row=r,column=1,value='措施项目费' if label == '安全施工费' else '').font=Font(name='微软雅黑', bold=True, size=10)
        ws.cell(row=r,column=3,value=label).font=df
        ws.cell(row=r,column=18,value=round(val,2)).font=df
        for c in range(1,len(headers)+1): ws.cell(row=r,column=c).border=bd

    r += 1
    social_fee = (total_sum + total_measures) * reg_rate
    ws.cell(row=r,column=1,value='规费').font=Font(name='微软雅黑', bold=True, size=10)
    ws.cell(row=r,column=3,value=f'规费({reg_rate*100:.2f}%)').font=df
    ws.cell(row=r,column=18,value=round(social_fee,2)).font=df
    for c in range(1,len(headers)+1): ws.cell(row=r,column=c).border=bd

    r += 1
    taxable = total_sum + total_measures + social_fee
    vat = taxable * vat_rate
    ws.cell(row=r,column=1,value='税金').font=Font(name='微软雅黑', bold=True, size=10)
    ws.cell(row=r,column=3,value=f'增值税({vat_rate*100:.2f}%)').font=df
    ws.cell(row=r,column=18,value=round(vat,2)).font=df
    for c in range(1,len(headers)+1): ws.cell(row=r,column=c).border=bd

    r += 1
    grand = taxable + vat
    ws.cell(row=r,column=1,value='总造价').font=Font(name='微软雅黑',bold=True,size=11)
    ws.cell(row=r,column=18,value=round(grand,2)).font=Font(name='微软雅黑',bold=True,size=11)
    for c in range(1,len(headers)+1): ws.cell(row=r,column=c).border=bd

    widths = [6,14,22,8,10,12,28,10,12,10,10,10,10,10,10,10,12,12,30]
    for idx,w in enumerate(widths,1):
        ws.column_dimensions[chr(64+idx) if idx<=26 else 'Z'].width = w

    xlsx_path = os.path.join(output_dir, '已组价清单.xlsx')
    wb.save(xlsx_path)

    # v5.16: 回填价格后的清单结果写回(供综合单价分析表等下游使用)
    try:
        with open(boq_json, 'w', encoding='utf-8') as f:
            json.dump(boq_items, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    print(f'  低置信度定额: {low_conf}/{len(boq_items)} 项')
    print(f'  分部分项费: {total_sum:,.2f}')
    print(f'  措施项目费: {total_measures:,.2f}')
    print(f'  规费: {social_fee:,.2f}')
    print(f'  增值税: {vat:,.2f}')
    print(f'  总造价: {grand:,.2f}')
    print(f'  输出: {xlsx_path}')
    return boq_items


if __name__ == '__main__':
    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'output')
    os.makedirs(out, exist_ok=True)
    sample = [{'seq':1,'code':'040101003','name':'细粒式沥青混凝土AC-13','unit':'m³','qty':31.38,'is_substitute':False}]
    with open(os.path.join(out,'清单结果.json'),'w', encoding='utf-8') as f: json.dump(sample,f,ensure_ascii=False)
    run(os.path.join(out,'清单结果.json'), out)
