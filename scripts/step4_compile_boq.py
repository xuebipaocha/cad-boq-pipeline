"""Step 4: 编清单 — v3.1 准确性增强版

新增能力：
- 清单匹配带置信度与候选列表。
- 输出推荐定额候选，供 Step 5 组价优先使用映射表。
- 低置信度项目明确标记为待复核。
- 单位换算：清单单位与算量单位不一致时自动换算。
- 清单质量报告：匹配率、低置信度项、单位偏差等。
- 分部自动归类：按专业和项目特征自动分配分部。
"""
import sys, os, json, re
sys.stdout.reconfigure(encoding='utf-8')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)
DATA_DIR = os.path.join(BASE_DIR, '..', 'data')
TEMPLATE = os.path.join(DATA_DIR, 'templates', '国标清单模板.xlsx')

SPEC_CAT = {
    '房屋建筑与装饰工程': '房屋建筑与装饰工程',
    '安装工程': '通用安装工程',
    '市政工程': '市政工程',
    '园林绿化工程': '园林绿化工程',
    '钢结构工程': '房屋建筑与装饰工程',
}

# 分部关键词映射 — 按专业自动归类
SECTION_MAP = {
    '房屋建筑与装饰工程': [
        ('土石方工程', ['挖','填','土方','石方','基坑','沟槽','回填','弃置']),
        ('地基与基础工程', ['基础','桩','承台','垫层','地基']),
        ('砌筑工程', ['砌','砖','砌体','墙']),
        ('混凝土及钢筋混凝土工程', ['混凝土','砼','钢筋','模板','柱','梁','板','楼梯']),
        ('门窗工程', ['门','窗','玻璃']),
        ('屋面及防水工程', ['防水','屋面','保温','隔热','涂膜']),
        ('楼地面装饰工程', ['楼地面','地坪','地板','石材','瓷砖']),
        ('墙柱面装饰工程', ['墙面','抹灰','涂料','裱糊','饰面']),
        ('天棚工程', ['天棚','吊顶']),
        ('措施项目', ['脚手架','模板','垂直运输','超高']),
    ],
    '市政工程': [
        ('道路工程', ['道路','路面','路基','基层','面层','侧石','路缘石','人行道']),
        ('桥涵工程', ['桥梁','涵洞','墩台','支座']),
        ('管网工程', ['管道','管网','检查井','雨水','污水','给水']),
        ('土石方工程', ['挖','填','土方','石方','沟槽','回填']),
    ],
    '安装工程': [
        ('给排水工程', ['给水','排水','管道','阀门','水表','卫生器具']),
        ('电气工程', ['电气','电缆','桥架','配管','配电箱','灯具','开关','插座']),
        ('消防工程', ['消防','喷淋','消火栓','报警','防火']),
        ('通风空调工程', ['通风','空调','风管','风机','盘管']),
    ],
    '园林绿化工程': [
        ('绿化工程', ['乔木','灌木','草坪','苗木','种植','绿地']),
        ('园路工程', ['铺装','路缘石','园路','步道']),
        ('园林设施', ['座椅','花坛','雕塑','景观']),
    ],
    '钢结构工程': [
        ('钢构件制作安装', ['钢梁','钢柱','H型','钢板','檩条','支撑']),
        ('涂装工程', ['防腐','防火','涂装','油漆']),
    ],
}


def _auto_section(name, specialty):
    """根据项目名称和专业自动分配分部。"""
    sections = SECTION_MAP.get(specialty, [])
    for section_name, keywords in sections:
        if any(kw in name for kw in keywords):
            return section_name
    return '其他工程'


def _clean_spec(spec):
    """规格清洗: 'DN=DN100 POWER=7.5kW' → 'DN100 7.5kW'(属性标签丢弃, 保留值)。
    占位值('未标管径'/'见表'/'无')返回空串 — 不进特征"""
    parts = []
    for tok in (spec or '').split():
        if '=' in tok:
            tok = tok.split('=', 1)[1]
        if tok and tok not in ('未标管径', '未标', '见表', '无', 'None'):
            parts.append(tok)
    return ' '.join(parts)


def _extract_project_name(recog):
    """v6.4: 从施工说明提取工程名称(如 '工程名称一工场船体大楼大修项目 耐火等级' → 
    '一工场船体大楼大修项目')。无 → 空串(模板示例文本保留原样)。
    """
    stop = ('耐火', '建设地点', '抗震', '结构形式', '建筑使用', '工程等级',
            '建设单位', '建筑面积', '设计单位', '序号', '图号', '页数')
    for note in (recog or {}).get('施工说明') or []:
        m = re.search(r'工程名称[：:]?\s*([^\s；;，,]{2,40})', str(note))
        if not m:
            continue
        name = m.group(1)
        for s in stop:
            idx = name.find(s)
            if idx > 0:
                name = name[:idx]
        name = name.strip()
        if 4 <= len(name) <= 30:
            return name
    return ''


# v5.17: 分项名 → 图纸特征匹配关键词(构造层/做法表/施工说明)
_ITEM_FEATURE_KEYWORDS = {
    '混凝土': ['混凝土', '砼'],
    '垫层': ['垫层'],
    '找平': ['找平'],
    '楼地面': ['楼地面', '地面'],
    '墙面': ['墙面', '墙'],
    '天棚': ['天棚', '吊顶'],
    '乳胶漆': ['乳胶漆', '涂料'],
    '砌体': ['砌体', '砌块', '砖'],
    '钢筋': ['钢筋', 'Φ', 'φ', '箍筋'],
    '模板': ['模板'],
    '土方': ['挖', '土方'],
    '回填': ['回填', '填'],
    '沥青': ['沥青'],
    '稳定': ['稳定', '水稳'],
    '碎石': ['碎石', '级配'],
    '透层': ['透层'],
    '防水': ['防水'],
    '保温': ['保温', '绝热'],
    '涂料': ['涂料', '防火'],
    '管道': ['管'],
    '阀门': ['阀'],
    '灯具': ['灯'],
    '桥架': ['桥架'],
    '电缆': ['电缆', '线缆'],
}


# v5.17: 自补清单编码 — 专业码 + 'B' + 3位序号(真实工程规则: 01B001/03B001/03B007)
_SPEC_PREFIX = {
    '房屋建筑与装饰工程': '01',
    '建筑与装饰工程': '01',
    '安装工程': '03',
    '市政工程': '04',
    '园林绿化工程': '05',
    '钢结构工程': '06',
    '给排水工程': '03',
    '电气工程': '03',
    '消防工程': '03',
}


def _make_supplement_code(specialty, name, seq):
    """生成自补清单编码: 专业前缀+B+序号(如 03B001)。国标清单无合适项时自补。"""
    prefix = _SPEC_PREFIX.get(specialty, '01')
    return f'{prefix}B{seq:03d}'


def _match_features_for_item(name, recog):
    """v5.17: 按分项名匹配图纸中相关的做法特征(构造层/施工说明)。

    返回如 '细石混凝土楼地面 40mm C20细石砼；水泥砂浆找平层 20mm 1:3水泥砂浆'
    (仅该分项相关的做法, 而非全图)。无匹配返回空串。
    """
    hits = []
    # 分项名关键词(取最长匹配优先)
    kws = []
    for kw, group in _ITEM_FEATURE_KEYWORDS.items():
        if kw in name or any(g in name for g in group):
            kws.extend(group)
    # v6.4: 并入分项名材料词 — 使做法表构造层(如'木质地板')可被匹配
    for w in ('木质地板', 'PVC', '橡胶', '塑胶', '无机涂料', '面砖', '防潮板', '浅黄色', '地砖', '地板'):
        if w in name:
            kws.append(w)
    if not kws:
        return ''
    # 构造层: 名称/材料含关键词
    for layer in (recog or {}).get('构造层') or []:
        text = ' '.join(str(layer.get(k) or '') for k in ('名称', '材料', '部位'))
        if any(k in text for k in kws):
            seg = []
            for k in ('名称', '材料', '部位'):
                v = str(layer.get(k) or '').strip()
                if v and v not in ('无', 'None'):
                    seg.append(v)
            if layer.get('厚度_mm'):
                seg.append(f"{layer['厚度_mm']}mm")
            if seg:
                hits.append(' '.join(seg))
    # 施工说明: 含关键词的"做法行"(v6.4: 需含分项材料词+做法特征, 防止碎片行乱入)
    # 原实现仅按关键词(如'墙')匹配, 图纸文字碎片化时会把无关行拼进特征
    _MAT_KW = ['抹灰', '砂浆', '防水', '涂料', '抗裂', '网格布', '腻子', '漆',
               '保温', '石材', '砖', '混凝土', '找平', '垫层', '面层', '乳胶漆',
               '脚手架', '卷材', '水泥', '细石', '垫层', '地板', '地砖', '面砖']
    _RECIPE_MARK = ['厚', '道', '遍', 'mm', '铺', '抹', '刷', '涂', '贴', '做',
                    '层', '砂浆', '防水', '涂料', '腻子', '网格布', '卷材', '面层']
    mat_kws = [w for w in _MAT_KW if w in name]
    # v6.4: 部位约束 — 分项名含部位词时, 匹配行必须提及该部位(防跨部位误配:
    # 如"外墙防水"不得引用"卫生间防水"做法)
    _PART_KW = ['外墙', '内墙', '屋面', '地面', '楼面', '楼梯', '天棚', '顶棚',
                '走廊', '卫生间', '室外', '室内', '立面']
    part_kws = [w for w in _PART_KW if w in name]
    for note in (recog or {}).get('施工说明') or []:
        s = str(note).strip()
        if not s or not any(k in s for k in kws):
            continue
        if mat_kws and not any(m in s for m in mat_kws):
            continue  # 行未提及该分项的材料 → 不相关碎片
        if part_kws and not any(p in s for p in part_kws):
            continue  # 行未提及该分项的部位 → 跨部位误配
        if not any(m in s for m in _RECIPE_MARK) or len(s) < 12 or len(s) > 160:
            continue  # 非做法行 / 过短碎片 / 超长拼接行
        if any(x in s for x in ('规程', '规范', '标准', 'GB5', 'GB/T', '图集', '建筑高度', '耐火等级', '工程名称', '页数', '抗震设防', '结构形式', '建设单位')):
            continue  # 规范引用/工程概况行不是做法
        hits.append(s)
    # 去重保序
    seen, out = set(), []
    for h in hits:
        if h not in seen:
            seen.add(h)
            out.append(h)
    return '；'.join(out[:4])


def build_features_map(cm):
    """v5.2: 构件模型规格 → 清单项目特征映射。

    构件编号与算量分项名的关联: 编号本身 / 编号+安装 / 编号+制作安装
    (如 阀门 → '阀门安装', PUMP → 'PUMP安装', 给水_DN100 → '给水_DN100')。
    无规格的构件不生成映射(回退分项名)。
    """
    feat = {}
    for cls, comps in (cm or {}).items():
        for c in comps or []:
            name = c.get('编号', '')
            spec = _clean_spec(c.get('规格', ''))
            if not name or not spec:
                continue
            for key in (name, f'{name}安装', f'{name}制作安装', f'{name}敷设'):
                feat[key] = f'{name}: 规格 {spec}'
    return feat


def _append_estimated_block(xlsx_path, estimated_items):
    """v6.6: 计价表尾部追加"待核实"区块 — 估算分项(量未核实)独立展示,
    不进正式清单量。"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    wb = load_workbook(xlsx_path)
    ws = wb.active
    r = ws.max_row + 2
    ws.cell(row=r, column=2, value='▼ 待核实分项（估算值，不编造 — 图纸无精确证据，需人工核实后回填正式清单）') \
        .font = Font(name='微软雅黑', bold=True, size=10, color='C00000')
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
    r += 1
    for ei in estimated_items:
        ws.cell(row=r, column=2, value=ei['分项名称']).font = Font(name='微软雅黑', size=10)
        ws.cell(row=r, column=4, value=ei['单位']).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5, value=ei['工程量']).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=10, value=str(ei['计算式'])[:60]).font = Font(name='微软雅黑', size=10)
        r += 1
    wb.save(xlsx_path)


def build_features_text(recog):
    """v5.17: 图纸做法/构造层/施工说明 → 项目特征文本(套定额依据)。

    合并: 构件规格映射 + 构造层(名称/材料/厚度/部位) + 施工说明 + 表格。
    供: ①step4 清单项目特征描述 ②step5 定额子目条件判定。
    """
    parts = []
    # 构件规格(安装 DN/功率/型号 等)
    try:
        fm = build_features_map((recog or {}).get('构件模型', {}))
        parts.extend(v for v in fm.values() if v)
    except Exception:
        pass
    # 构造层(做法表提取): "细石混凝土楼地面 40mm C20细石砼"
    for layer in (recog or {}).get('构造层') or []:
        seg = []
        for k in ('名称', '材料', '部位'):
            v = str(layer.get(k) or '').strip()
            if v and v not in ('无', 'None'):
                seg.append(v)
        if layer.get('厚度_mm'):
            seg.append(f"{layer['厚度_mm']}mm")
        if seg:
            parts.append(' '.join(seg))
    # 施工说明(做法表原文)
    for note in (recog or {}).get('施工说明') or []:
        s = str(note).strip()
        if s:
            parts.append(s)
    # 表格(做法表结构化)
    for tbl in (recog or {}).get('表格') or []:
        s = str(tbl).strip()
        if s:
            parts.append(s)
    # v6.2: 设计说明材料规格(如 800×800/300×600/C30/DN25) → 并入特征文本(供主材匹配/套定额)
    dn = (recog or {}).get('设计说明') or {}
    for spec in dn.get('材料规格') or []:
        if spec.get('规格'):
            parts.append(f"材料规格{spec['规格']}")
    return '；'.join(parts)


def _convert_qty(qty, from_unit, to_unit):
    """单位换算：清单单位与算量单位不一致时转换工程量。"""
    if not from_unit or not to_unit or from_unit == to_unit:
        return qty, 1.0, ''
    try:
        from pipeline.db import quota_qty_from_list_qty
        converted, factor, note = quota_qty_from_list_qty(qty, from_unit, to_unit)
        return round(converted, 4), round(factor, 4), note
    except Exception:
        # v4.0: 换算失败不再静默——返回原始值并标记
        return qty, 1.0, '单位换算失败，未换算'


def compile_boq_quality_report(boq_items):
    """生成清单编制质量报告。v5.6: 增加数据来源可信度统计"""
    warnings = []
    total = len(boq_items)
    if total == 0:
        return {'质量分': 0, '警告数量': 1, '警告': [{'级别':'高','问题':'清单为空','建议':'检查上游算量结果'}]}

    matched = sum(1 for i in boq_items if not i.get('is_substitute'))
    match_rate = matched / total
    low_conf = sum(1 for i in boq_items if i.get('match_confidence') in ('低', '待确认'))
    high_conf = sum(1 for i in boq_items if i.get('match_confidence') == '高')
    unit_mismatch = sum(1 for i in boq_items if i.get('unit_note') and '不一致' in i.get('unit_note', ''))
    # v5.6: 数据来源可信度
    est_n = sum(1 for i in boq_items if i.get('data_source') == '估算')
    pending_n = sum(1 for i in boq_items if i.get('data_source') == '待提取')

    if match_rate < 0.5:
        warnings.append({'级别':'高','问题':f'清单匹配率仅{match_rate:.0%}（{matched}/{total}）','建议':'检查分项名称规范性和数据库覆盖度'})
    elif match_rate < 0.8:
        warnings.append({'级别':'中','问题':f'清单匹配率{match_rate:.0%}，{total-matched}项待匹配','建议':'人工复核待匹配项'})

    if low_conf > total * 0.3:
        warnings.append({'级别':'中','问题':f'低置信度项{low_conf}个，占{low_conf/total:.0%}','建议':'补充清单映射表或修正分项名称'})

    if unit_mismatch > 0:
        warnings.append({'级别':'低','问题':f'{unit_mismatch}项单位维度不一致','建议':'检查算量单位与清单单位是否匹配'})

    if high_conf / total < 0.3 and total > 3:
        warnings.append({'级别':'低','问题':f'高置信度匹配仅{high_conf}项','建议':'扩充领域词典或映射表'})

    # v5.6: 估算/待提取占比过高 → 数据可信度警告
    if est_n + pending_n > 0:
        ratio = (est_n + pending_n) / total
        if ratio > 0.3:
            warnings.append({'级别':'高','问题':f'估算/待提取值占比 {ratio:.0%}（{est_n+pending_n}/{total}），数据可信度低','建议':'补齐图纸参数后重算'})
        elif ratio > 0.1:
            warnings.append({'级别':'中','问题':f'估算/待提取值 {est_n+pending_n} 项，占{ratio:.0%}','建议':'重点复核估算项'})

    # v6.5: 合计加和性卡口 — 合价合计 = 明细之和(防止 149≠134 类错误)
    # 未匹配编码的明细项独立列示, 不得被合计吞掉
    def _is_total_row(i):
        return '合计' in str(i.get('name', '')) or '总计' in str(i.get('name', ''))
    priced = [i for i in boq_items if i.get('合价') is not None and not _is_total_row(i)]
    if priced:
        sum_parts = round(sum(float(i.get('合价', 0) or 0) for i in priced), 2)
        total_price = None
        # 合计可能在最后一项(分部汇总行)或独立 total 字段
        for i in boq_items:
            if _is_total_row(i):
                total_price = i.get('合价')
                break
        if total_price is not None:
            diff = round(abs(float(total_price) - sum_parts), 2)
            if diff > 0.01:
                warnings.append({'级别': '高',
                                 '问题': f'合计加和性错误: 明细合价之和 {sum_parts} ≠ 合计 {total_price} (差 {diff})',
                                 '建议': '检查分部汇总行与明细项是否遗漏/重复'})
        # 未匹配编码明细独立列示检查
        unmapped = [i for i in priced if i.get('is_substitute') and i.get('合价')]
        if unmapped:
            unmapped_sum = round(sum(float(i.get('合价', 0) or 0) for i in unmapped), 2)
            warnings.append({'级别': '低',
                             '问题': f'自补清单项 {len(unmapped)} 项合价 {unmapped_sum} 元(独立列示, 未并入标准项合计)',
                             '建议': '自补项需人工确认单价后并入总价'})

    score = max(0, 100 - sum({'高':20,'中':10,'低':4}.get(w['级别'],5) for w in warnings))
    score = int(score * (0.5 + 0.5 * match_rate))  # 匹配率权重
    return {
        '质量分': score,
        '匹配率': f'{match_rate:.0%} ({matched}/{total})',
        '高置信度': high_conf,
        '低置信度': low_conf,
        '待匹配': total - matched,
        '单位不一致': unit_mismatch,
        '估算值': est_n,
        '待提取': pending_n,
        '警告数量': len(warnings),
        '警告': warnings[:50],
    }


def run(calc_json, output_dir):
    print('='*50)
    print('Step 4: 编清单 — v3.1 增强匹配')
    print('='*50)
    with open(calc_json, 'r', encoding='utf-8') as f:
        calc_results = json.load(f)

    specialty = '房屋建筑与装饰工程'
    recog_path = os.path.join(output_dir, '识图结果.json')
    if os.path.exists(recog_path):
        with open(recog_path, 'r', encoding='utf-8') as f:
            recog = json.load(f)
            specialty = recog.get('专业类型', '房屋建筑与装饰工程')

    print(f'  专业: {specialty}')
    from pipeline.db import find_list_item, get_mapped_quotas, infer_unit

    # v5.2: 构件模型规格 → 清单项目特征(安装 DN/功率/型号 等)
    feat_map = {}
    recog_data = {}
    features_text = ''
    try:
        if os.path.exists(recog_path):
            with open(recog_path, 'r', encoding='utf-8') as f:
                recog_data = json.load(f)
                feat_map = build_features_map(recog_data.get('构件模型', {}))
                # v5.17: 完整特征文本 = 构件规格 + 构造层 + 施工说明 + 表格(套定额依据)
                features_text = build_features_text(recog_data)
    except Exception:
        feat_map = {}
        features_text = ''

    cat = SPEC_CAT.get(specialty, specialty)
    boq_items = []
    estimated_items = []  # v6.6: 估算分项(数据来源=估算)独立分流, 不进正式清单
    skipped = 0
    supplement_seq = 1  # v5.17: 自补清单序号
    pending_items = []  # v5.6: 待提取分项独立交付, 不进清单
    project_name = ''   # v6.4: 工程名称(图纸提取, 覆盖模板示例)
    for item in calc_results:
        name = item.get('分项名称', '')
        qty = item.get('工程量', 0) or 0
        # v6.3 C1: 施工范围外分项不进清单(标记范围外)
        if item.get('范围外'):
            skipped += 1
            continue
        # v5.6: 数据来源卡口 — 待提取分项不进清单, 进待提取清单
        src = item.get('数据来源', '')
        if src == '待提取' or ('待CAD提取' in item.get('计算式','') or '待输入' in item.get('计算式','')):
            if src == '待提取':
                pending_items.append({
                    '分项名称': name, '单位': item.get('单位', ''),
                    '原因': item.get('计算式', '无证据'), '数据来源': '待提取',
                })
                continue
            if qty <= 0:
                skipped += 1
                continue
        calc_unit = infer_unit(name, item.get('单位', ''))
        section = _auto_section(name, specialty)

        # 清单匹配
        lst = find_list_item(name, category=cat, top_n=5)
        best = lst[0] if lst else {}
        score = best.get('_score', 0) or 0
        # v6.4: 房建专业拆除分项无合适国标项(拆除类多为市政/爆破), 强制自补防错配(如'楼梯墙拆除'→'楼梯')
        # v6.6: 大修泛化分项(门窗更换/雨水管更换/洞口面积)同样无对应国标科目 —
        # '门窗更换'错配'门窗框槛'(仿古, 单位樘vs个)比自补更糟, 强制自补更诚实
        if ('拆除' in name or '更换' in name or name == '门窗洞口面积') \
                and specialty in ('房屋建筑与装饰工程', '建筑与装饰工程'):
            best, score = {}, 0.0
        # v4.0: 匹配确认阈值 0.35→0.5, 且低置信度(0.25-0.45)必须标记待复核
        # v5.17: 低置信度不标"待匹配"——按国标自补清单规则生成 B 类编码(专业码+B+序号)
        confident = bool(best) and score >= 0.45
        if confident:
            code = best.get('item_code')
            list_name = best.get('item_name')
        else:
            code = _make_supplement_code(specialty, name, supplement_seq)
            supplement_seq += 1
            list_name = name
        list_unit = infer_unit(list_name, best.get('unit') if confident else calc_unit)

        # 单位换算
        final_qty, conv_factor, unit_note = _convert_qty(qty, calc_unit, list_unit)

        # 定额映射(自补清单无国标映射, 走关键词检索)
        mapped = get_mapped_quotas(code, top_n=5) if confident else []

        # v5.6: 估算分项强制标记(数据来源=估算), 统计单独走
        is_estimated = (item.get('数据来源') == '估算')
        if confident:
            review_note = ''
        elif 'B' in (code or '') and code[:2].isdigit():
            review_note = '国标清单无合适项，自补清单'
        else:
            review_note = '清单低置信度匹配，需人工复核'
        if is_estimated:
            review_note = (review_note + '；' if review_note else '') + '估算值，需核实'

        # v5.17: 项目特征 = 构件规格 + 与该分项相关的构造层/做法表/施工说明
        # (特征描述必须根据图纸完善, 供套定额判定)
        feat_spec = feat_map.get(name, '')
        feat_layers = _match_features_for_item(name, recog_data)
        if feat_spec and feat_layers:
            features = f'{feat_spec}；{feat_layers}'
        else:
            features = feat_spec or feat_layers or name

        # v6.4: 工程名称(图纸施工说明提取, 覆盖模板示例文本)
        if not project_name:
            project_name = _extract_project_name(recog_data)

        # v6.6: 估算分项质量卡口 — 数据来源=估算 不进正式清单(量未核实),
        # 分流到独立"待核实清单"交付(可见/不编造/核实后人工回填)
        if is_estimated:
            estimated_items.append({
                '分项名称': name, '单位': calc_unit, '工程量': qty,
                '计算式': item.get('计算式', ''), '数据来源': '估算',
                '备注': item.get('备注', ''),
            })
            continue

        boq_items.append({
            'seq': len(boq_items) + 1,
            'code': code,
            'name': list_name or name,
            'source_name': name,
            'section': section,
            'project': project_name,
            'unit': list_unit or calc_unit,
            'calc_unit': calc_unit,
            'qty': final_qty,
            'original_qty': qty,
            'conv_factor': conv_factor,
            'unit_note': unit_note,
            # v5.2/v5.17: 项目特征优先取构件模型规格, 回退分项名
            'features': features,
            'features_text': features_text,  # v5.17: 全图特征文本(套定额条件判定)
            'is_substitute': not confident,
            'match_score': round(score, 4),
            'match_confidence': best.get('_confidence', '待确认') if best else '待确认',
            'match_method': best.get('_match_method', 'none') if best else 'none',
            'list_candidates': lst[:5],
            'mapped_quotas': mapped[:5],
            'review_note': review_note,
            'data_source': item.get('数据来源', '实测'),  # v5.6 数据来源四级
        })

    # 输出Excel
    if os.path.exists(TEMPLATE):
        from pipeline.fill_boq import fill_boq_template
        xlsx = os.path.join(output_dir, '分部分项工程量清单计价表.xlsx')
        fill_boq_template(boq_items, TEMPLATE, xlsx)
        if estimated_items:
            _append_estimated_block(xlsx, estimated_items)
    else:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        wb = Workbook(); ws = wb.active
        hf = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        hft = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF'); df = Font(name='微软雅黑', size=10)
        bd = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        ws.cell(row=1,column=1,value='分部分项工程量清单计价表').font = Font(name='微软雅黑', bold=True, size=16)
        ws.merge_cells('A1:I1'); ws.cell(row=1,column=1).alignment = Alignment(horizontal='center')
        headers = ['序号','项目编码','项目名称','单位','工程量','分部','匹配分','置信度','备注']
        for i,h in enumerate(headers,1):
            c=ws.cell(row=3,column=i,value=h); c.font=hft; c.fill=hf; c.alignment=Alignment(horizontal='center'); c.border=bd
        for ri,item in enumerate(boq_items):
            r=ri+4
            vals=[item['seq'],item['code'],item['name'],item['unit'],item['qty'],
                  item.get('section',''),item['match_score'],item['match_confidence'],item.get('review_note','')]
            for ci,v in enumerate(vals,1):
                ws.cell(row=r,column=ci,value=v).font=df; ws.cell(row=r,column=ci).border=bd
                ws.cell(row=r,column=ci).alignment=Alignment(horizontal='center', wrap_text=True)
        if estimated_items:
            r += 1
            ws.cell(row=r,column=2,value='▼ 待核实分项（估算值，不编造 — 图纸无精确证据，需人工核实后回填正式清单）').font = Font(name='微软雅黑', bold=True, size=10, color='C00000')
            ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9)
            r += 1
            for ei in estimated_items:
                ws.cell(row=r,column=2,value=ei['分项名称']).font=df
                ws.cell(row=r,column=4,value=ei['单位']).font=df
                ws.cell(row=r,column=5,value=ei['工程量']).font=df
                ws.cell(row=r,column=9,value=ei['计算式'][:60]).font=df
                r += 1
        for c,w in [('A',6),('B',16),('C',28),('D',8),('E',10),('F',14),('G',10),('H',10),('I',24)]:
            ws.column_dimensions[c].width = w
        xlsx = os.path.join(output_dir, '分部分项工程量清单计价表.xlsx')
        wb.save(xlsx)

    # 质量报告
    report = compile_boq_quality_report(boq_items)
    matched = sum(1 for i in boq_items if not i.get('is_substitute'))
    if skipped:
        print(f'  ⚠ 跳过 {skipped} 个 0 量占位项（待CAD提取）')
    # v5.6: 数据来源统计
    est_n = len(estimated_items)
    pending_n = len(pending_items)
    if est_n or pending_n:
        print(f'  ⚠ 估算值 {est_n} 项（分流待核实清单）| 待提取 {pending_n} 项（已拦截不进清单）')
    print(f'  清单匹配: {matched}/{len(boq_items)} 项 | 质量分: {report["质量分"]} | 警告: {report["警告数量"]}条')
    print(f'  输出: {xlsx}')
    json_path = os.path.join(output_dir, '清单结果.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(boq_items, f, ensure_ascii=False, indent=2)
    # v5.6: 待提取清单独立交付
    if pending_items:
        pend_path = os.path.join(output_dir, '待提取清单.json')
        with open(pend_path, 'w', encoding='utf-8') as f:
            json.dump(pending_items, f, ensure_ascii=False, indent=2)
        print(f'  待提取清单: {pend_path}')
    # v6.6: 待核实清单独立交付(估算分项 — 不进正式清单量, 交付可见)
    if estimated_items:
        est_path = os.path.join(output_dir, '待核实清单.json')
        with open(est_path, 'w', encoding='utf-8') as f:
            json.dump(estimated_items, f, ensure_ascii=False, indent=2)
        print(f'  待核实清单: {est_path}')
    with open(os.path.join(output_dir, '清单质量报告.json'), 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    return boq_items


if __name__ == '__main__':
    out = os.path.join(os.path.dirname(BASE_DIR), 'output')
    os.makedirs(out, exist_ok=True)
    recog_json = os.path.join(out, '识图结果.json')
    calc_json = os.path.join(out, '算量结果.json')
    with open(recog_json, 'w', encoding='utf-8') as f:
        json.dump({'专业类型':'市政工程'}, f, ensure_ascii=False)
    with open(calc_json, 'w', encoding='utf-8') as f:
        json.dump([{'分项名称':'沥青混凝土','工程量':31.38,'单位':'m³'},
                   {'分项名称':'水泥稳定碎石','工程量':15.5,'单位':'m³'},
                   {'分项名称':'侧石','工程量':120,'单位':'m'}], f, ensure_ascii=False)
    run(calc_json, out)
