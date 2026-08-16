"""
Step 3: 算量 — v3.1 增强版

改进（相比 v3.0）：
- 新增参数化公式引擎：支持从识图数据中提取参数（厚度、挖深、层高、构件数量），替代固定系数。
- 新增规则引擎与专业计算器交叉校验：同一分项两种方法都算，偏差超阈值告警。
- 体积计算优先使用构造层厚度，不再固定 0.5m。
- 长度类公式按线性构件类型筛选求和。
- count 公式使用实际图块/苗木数量，不再硬编码 10。
- 输出算量质量报告，提示估算项、缺厚度、缺线性构件和异常量。
"""
import sys, os, json, sqlite3, re
sys.stdout.reconfigure(encoding='utf-8')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)


def _areas(data):
    return sum(a.get('面积_m2',0) or 0 for a in data.get('面积区域',[]))


def _layer_thickness(data, layer_name=None):
    """提取构造层厚度(mm)，可按名称筛选。返回平均厚度(m)或 None。"""
    layers = data.get('构造层', [])
    if layer_name:
        layers = [l for l in layers if layer_name in (l.get('名称','') + l.get('材料',''))]
    vals = [l.get('厚度_mm') for l in layers if isinstance(l.get('厚度_mm'), (int, float)) and l.get('厚度_mm') > 0]
    return (sum(vals) / len(vals) / 1000) if vals else None


def _total_thickness_m(data):
    """所有构造层厚度之和(m)，用于多层结构总厚度。"""
    vals = [l.get('厚度_mm') for l in data.get('构造层',[]) if isinstance(l.get('厚度_mm'), (int,float)) and l.get('厚度_mm') > 0]
    return sum(vals) / 1000 if vals else None


def _linear_len(data, keyword=None):
    items = data.get('线性构件', [])
    if keyword:
        items = [i for i in items if keyword in (i.get('名称','') + i.get('类型','') + i.get('系统',''))]
    return sum(i.get('长度_m',0) or 0 for i in items)


def _count_blocks(data, block_type=None):
    """从图块统计中获取数量。"""
    blocks = data.get('图块') or {}
    if block_type:
        d = blocks.get(block_type, {})
        return sum(d.values()) if isinstance(d, dict) else 0
    total = 0
    for key in ['tree_blocks','equip_blocks','valve_blocks','light_blocks','switch_blocks','panel_blocks','sanitary_blocks','fire_blocks','hardscape_blocks']:
        d = blocks.get(key, {}) or {}
        total += sum(d.values()) if isinstance(d, dict) else 0
    return total


def _count_from_info(data, *paths):
    """从安装信息/园林信息中按路径取数量。"""
    for path in paths:
        obj = data
        try:
            for p in path:
                obj = obj.get(p, {})
            if isinstance(obj, list):
                return sum(i.get('数量',0) for i in obj)
            if isinstance(obj, dict):
                return sum(obj.values())
        except:
            pass
    return 0


def calculate_by_formula(fc, data, rule_text=''):
    """根据公式代码和数据计算工程量。v3.1 支持参数化提取。"""
    area = _areas(data)

    if fc == 'calc_area':
        return area, '面积合计'

    if fc == 'calc_vol':
        # 优先使用构造层厚度
        thick = _layer_thickness(data)
        if thick:
            return area * thick, f'面积×厚度{thick:.3f}m'
        # 多层结构总厚度
        total_thick = _total_thickness_m(data)
        if total_thick:
            return area * total_thick, f'面积×构造层总厚{total_thick:.3f}m'
        return area * 0.5, '面积×默认厚度0.5m（估算）'

    if fc == 'calc_len':
        # 按规则文本中的关键词筛选线性构件
        for kw in ['管道','管线','侧石','路缘石','边石','栏杆','墙线','桥架']:
            if kw in rule_text:
                length = _linear_len(data, kw)
                if length > 0:
                    return length, f'{kw}长度合计'
        return _linear_len(data), '线性构件长度合计'

    if fc == 'calc_count':
        # 优先从图块统计获取数量
        count = _count_blocks(data)
        if count > 0:
            return count, '图块/设备/苗木数量合计'
        # 从园林信息获取
        tree_count = _count_from_info(data, ['园林信息','苗木','乔木'])
        if tree_count > 0:
            return tree_count, '苗木数量合计'
        return 0, '未找到可计数构件'

    if fc == 'calc_rebar':
        # 面积 × 含钢量
        kg_m2 = 65
        # 从施工说明中尝试提取含钢量
        texts = ' '.join(data.get('施工说明', []))
        m = re.search(r'(\d+(?:\.\d+)?)\s*kg/m[2²]', texts)
        if m:
            kg_m2 = float(m.group(1))
        return area * kg_m2 / 1000, f'面积×{kg_m2}kg/m²含钢量（估算）'

    return 0, '未知公式'


def _apply_scope_mask(drawing_data, results):
    """v6.5: 施工范围掩码 — 设计内容(设计意图范围) + 不含项 裁剪算量分项。
    范围外分项标记'范围外', 不进清单。"""
    try:
        intent = (drawing_data.get('设计意图') or {})
        excludes = (intent.get('算量边界') or {}).get('不含项') or []
        design_scope = (intent.get('设计内容') or []) or []
        if excludes or design_scope:
            scope_parts = set()
            for it in design_scope:
                p = it.get('部位') or ''
                o = it.get('对象') or ''
                if p:
                    scope_parts.add(p)
                if o and o != '墙':
                    scope_parts.add(o)
            for it in results:
                nm = it.get('分项名称', '')
                out_reason = ''
                # 1) 不含项显式排除
                for ex in excludes:
                    if ex and ex.strip() and ex.strip() in nm:
                        out_reason = f'施工范围不含[{ex.strip()}]'
                        break
                # 2) 大修有设计内容但分项不在范围内 → 范围外(待确认)
                if not out_reason and scope_parts:
                    # v6.9.3: '门窗'对象覆盖所有门/窗分项(钢质门更换/塑钢窗更换…
                    # 材质拆项后名称含'门'或'窗'但不再含'门窗'连续词)
                    in_scope = any(p in nm for p in scope_parts if p)
                    if not in_scope and '门窗' in scope_parts and any(k in nm for k in ('门', '窗')):
                        in_scope = True
                    if not in_scope:
                        out_reason = '设计内容未覆盖该分项'
                if out_reason:
                    it['范围外'] = True
                    it.setdefault('备注', '')
                    it['备注'] = (it['备注'] + '；' if it['备注'] else '') + out_reason
    except Exception:
        pass
    return results


def calculate(drawing_data):
    specialty = drawing_data.get('专业类型', '房屋建筑与装饰工程')
    # v5.4 D6: 大修/改造 + 房建 → 专用计算器, 跳过规则引擎与新建模板
    # (规则引擎是"关键词驱动"的新建分项, 大修图纸会误触发土方/场地平整等)
    if specialty == '房屋建筑与装饰工程' and drawing_data.get('工程性质') == '大修与改造':
        try:
            from calc_renovation import calc as calc_renovation
            return _attach_basis(drawing_data, _apply_scope_mask(drawing_data, calc_renovation(drawing_data)))
        except Exception as e:
            print(f'  大修计算器跳过: {e}')
            # 回退到常规路径
    all_results = []
    used_items = set()

    db = os.path.join(os.path.dirname(BASE_DIR), 'data', 'liaoning_24.db')
    conn = sqlite3.connect(db); conn.row_factory = sqlite3.Row
    texts = ' '.join(drawing_data.get('施工说明', []))
    # v4.0: 规则引擎改为"说明关键词驱动 + 公式匹配", 并限制同公式同专业只触发一次
    extra_kws = [kw for kw in ['挖','填','砌','浇筑','安装','铺设','抹灰','涂刷','吊装','焊接','防腐','保温','防水','装修','拆除'] if kw in texts]

    if extra_kws:
        for kw in extra_kws:
            rules = conn.execute("""
                SELECT r.* FROM measure_rules r
                WHERE (r.category LIKE ? OR r.applicable_specialties LIKE ? OR r.applicable_specialties IS NULL)
                  AND r.rule_text LIKE ? AND r.formula_code IS NOT NULL
                ORDER BY r.id LIMIT 3
            """, (f'%{specialty[:4]}%', f'%{specialty}%', f'%{kw}%')).fetchall()
            for rule in rules:
                fc = rule['formula_code']
                name = (rule['output_name'] or rule['item_name'] or f'{kw}工程')[:30]
                qty, calc_note = calculate_by_formula(fc, drawing_data, rule['rule_text'] or '')
                if qty > 0 and name not in used_items:
                    used_items.add(name)
                    fu = rule['formula_unit'] or ''
                    rtext = rule['rule_text'] or ''
                    _rule_item = {
                        '分项名称': name, '单位': fu, '工程量': round(qty,2),
                        '计算式': calc_note, '定额编号': '',
                        '规则来源': f'第{rule["page_num"]}页', '备注': '规则引擎',
                        '_engine': 'rule'
                    }
                    _rule_item['数据来源'] = _infer_source(_rule_item)
                    all_results.append(_rule_item)
    conn.close()

    CALC_MAP = {
        '房屋建筑与装饰工程': 'calc_building',
        '安装工程': 'calc_mep',
        '市政工程': 'calc_civil',
        '园林绿化工程': 'calc_garden',
        '钢结构工程': 'calc_steel',
    }
    # v5.4 D6: 工程性质分发 — 大修/改造 + 房建 → calc_renovation(专用计算器)
    project_nature = drawing_data.get('工程性质', '新建')
    if specialty == '房屋建筑与装饰工程' and project_nature == '大修与改造':
        mod_name = 'calc_renovation'
    else:
        mod_name = CALC_MAP.get(specialty)
    if mod_name:
        try:
            mod = __import__(mod_name)
            calc_results = mod.calc(drawing_data)
            for item in calc_results:
                name = item.get('分项名称', '')
                if name not in used_items:
                    used_items.add(name)
                    item['_engine'] = 'specialty'
                    # v5.6: 数据来源缺省按引擎推断(大修计算器已自带来源标记)
                    if '数据来源' not in item:
                        item['数据来源'] = _infer_source(item)
                    all_results.append(item)
        except Exception as e:
            print(f'  专业计算器跳过: {e}')

    # v4.2: 表格→清单联动 (门窗表/桩表 → 分项)
    try:
        from table_to_boq import table_to_boq_items
        for item in table_to_boq_items(drawing_data.get('表格', [])):
            name = item.get('分项名称', '')
            if name not in used_items:
                used_items.add(name)
                item['_engine'] = 'table'
                if '数据来源' not in item:
                    item['数据来源'] = _infer_source(item)
                all_results.append(item)
    except Exception as e:
        print(f'  表格联动跳过: {e}')

    # v4.2: 剖面联动算量 (断面面积×平面长度 → 分项)
    try:
        for sq in drawing_data.get('剖面算量', []) or []:
            name = f'{sq.get("剖面", "剖面")}断面工程'
            if name not in used_items:
                used_items.add(name)
                _sq_item = {
                    '分项名称': name,
                    '单位': 'm³',
                    '工程量': sq.get('体积_m3', 0),
                    '计算式': f'断面面积{sq.get("断面面积_m2")}m²×平面长度{sq.get("平面长度_m")}m',
                    '定额编号': '',
                    '备注': '剖面联动',
                    '_engine': 'section',
                }
                _sq_item['数据来源'] = _infer_source(_sq_item)
                all_results.append(_sq_item)
    except Exception as e:
        print(f'  剖面联动跳过: {e}')
    return _attach_basis(drawing_data, all_results)


def _attach_basis(drawing_data, results):
    """v6.8: 分项依据引用 — 每个分项带计算规则依据(从知识库查),
    查不到标'待查证'(诚实: 有依据才写, 不编造依据)。
    """
    try:
        from knowledge_query import query_rule
        specialty = drawing_data.get('专业类型', '')
        # v6.9: 依据挂图纸证据 — 图名/图号+版本(版本意识: 每个数能对到哪版图纸)
        meta = drawing_data.get('图纸元数据', {}) or {}
        ver = drawing_data.get('图纸版本', {}) or {}
        doc_ref = meta.get('图名', '') or ''
        if ver.get('修订记录'):
            doc_ref += f"({ver['修订记录'][0]})"
        for it in results:
            if it.get('依据'):
                continue
            name = it.get('分项名称', '')
            # 用规则库关键词逐项匹配
            basis = ''
            for kw in ('防水', '混凝土', '门窗', '钢构件', '金属结构', '砌体',
                       '管道', '电缆', '苗木', '道路', '断面', '拆除', '抹灰',
                       '涂料', '楼地面', '地面', '墙面', '天棚', '吊顶', '地板', '地砖'):
                if kw in name:
                    r = query_rule(specialty, kw)
                    if r and r.get('规则'):
                        basis = f"{r.get('条款', '')}: {r['规则'][:60]}"
                        break
            it['依据'] = (basis or '待查证') + (f'；图纸: {doc_ref}' if doc_ref else '')
    except Exception:
        for it in results:
            it.setdefault('依据', '待查证')
    return results


def _infer_source(item):
    """v5.6: 数据来源推断 — 计算式含'估算'/'待提取'/'默认' → 估算/待提取, 否则实测"""
    from quality import SRC_MEASURED, SRC_ESTIMATED, SRC_PENDING
    note = str(item.get('计算式', '')) + str(item.get('备注', ''))
    if '待提取' in note or '待CAD提取' in note:
        return SRC_PENDING
    if '估算' in note or '默认' in note or '系数' in note:
        return SRC_ESTIMATED
    return SRC_MEASURED


def cross_validate(results):
    """规则引擎与专业计算器结果交叉校验 — v4.0: 规范化名称模糊对齐

    修复 v3.1 假交叉: 要求同名分项才比较 → 规则引擎(如'土方开挖')与专业计算器
    (如'挖一般土方')名称不同永远空转。现按去噪规范化名称对齐:
    去 '工程/安装/铺设' 等尾词 + 去单位/规格噪声后比较。
    """
    import re
    warnings = []
    def norm(name):
        n = re.sub(r'[（(].*?[)）]', '', name or '')
        n = re.sub(r'工程$|安装$|铺设$|敷设$|制作$|施工$|项目$', '', n)
        return n
    rule_items = {norm(r['分项名称']): r for r in results if r.get('_engine') == 'rule'}
    spec_items = {norm(r['分项名称']): r for r in results if r.get('_engine') == 'specialty'}
    for name in set(rule_items.keys()) & set(spec_items.keys()):
        r_qty = rule_items[name].get('工程量', 0)
        s_qty = spec_items[name].get('工程量', 0)
        if r_qty > 0 and s_qty > 0:
            err = abs(r_qty - s_qty) / max(r_qty, s_qty)
            if err > 0.3:
                warnings.append({
                    '级别': '中',
                    '问题': f'"{name}" 规则引擎({r_qty})与专业计算器({s_qty})偏差 {err:.0%}',
                    '建议': '检查计算参数和公式适用条件'
                })
    return warnings


def quality_report(drawing_data, results):
    warnings = cross_validate(results)
    area = _areas(drawing_data)
    # v6.7: 造价经验指标自检 — 单方含量超限 → 复核意见(像造价人的"心里有数")
    try:
        from indicators import calc_indicators, review_opinions
        drawing_data['造价指标'] = calc_indicators(drawing_data, results)
        for op in review_opinions(drawing_data, results):
            warnings.append({'级别': '中', '问题': op['意见'][:80],
                             '建议': f"计算式: {op['计算式']}；经验区间: {op['经验区间']}（{op['来源']}）"})
    except Exception:
        pass
    # v6.8: ABC 大项分析 — 按工程量排序列前10项+累计占比(审计反向思维:
    # 审价人先盯大项, 大项错=总价废; 大项自动标注重点复核)
    try:
        ranked = sorted([i for i in results if (i.get('工程量', 0) or 0) > 0],
                        key=lambda i: i.get('工程量', 0), reverse=True)
        if ranked:
            top = ranked[:10]
            abc = [{'分项名称': i.get('分项名称', ''), '单位': i.get('单位', ''),
                    '工程量': i.get('工程量', 0), '依据': i.get('依据', '')} for i in top]
            drawing_data['ABC大项'] = abc
            # 大项中"估算/待核"来源的 → 高风险警告(大项必须最严格)
            for i in top:
                src = i.get('数据来源', '')
                if src in ('估算', '待提取') or '待核' in str(i.get('计算式', '')):
                    warnings.append({'级别': '高', '问题':
                                     f'大项[{i.get("分项名称", "")}] 量{i.get("工程量", 0)}'
                                     f'{i.get("单位", "")} 来源{src or "待核"} — 大项必须实测/复核',
                                     '建议': '补图纸证据或人工核实后再进清单'})
    except Exception:
        pass
    if area <= 0:
        warnings.append({'级别':'高','问题':'缺少有效面积区域','建议':'检查闭合多段线或面积文字标注'})
    if not any(isinstance(l.get('厚度_mm'), (int,float)) and l.get('厚度_mm') > 0 for l in drawing_data.get('构造层',[])):
        warnings.append({'级别':'中','问题':'构造层厚度缺失','建议':'补充施工说明厚度或CAD标注关联'})
    if not drawing_data.get('线性构件'):
        warnings.append({'级别':'中','问题':'线性构件为空','建议':'检查管道、路缘石、墙线等图层识别'})
    for item in results:
        qty = item.get('工程量', 0) or 0
        name = item.get('分项名称','')
        unit = item.get('单位','')
        if qty <= 0:
            warnings.append({'级别':'高','问题':f'{name} 工程量为0','建议':'检查上游识图字段'})
        if area > 0 and unit in ('t','吨') and '钢筋' in name:
            kg_m2 = qty * 1000 / area
            if kg_m2 < 20 or kg_m2 > 120:
                warnings.append({'级别':'中','问题':f'钢筋含量 {kg_m2:.1f}kg/m² 异常','建议':'复核钢筋计算方式'})
        if '默认' in (item.get('计算式','') + item.get('备注','')) or '估算' in (item.get('计算式','') + item.get('备注','')):
            warnings.append({'级别':'低','问题':f'{name} 使用估算逻辑','建议':'补齐图纸参数后重算'})
    score = max(0, 100 - sum({'高':20,'中':10,'低':4}.get(w['级别'],5) for w in warnings))
    out = {'质量分': score, '警告数量': len(warnings), '警告': warnings[:100]}
    # v6.8: 自检结果随质量报告输出(供编制说明/下游消费)
    out['ABC大项'] = drawing_data.get('ABC大项', [])
    out['造价指标'] = drawing_data.get('造价指标', {})
    return out


def export_excel(results, path, pending=None):
    """导出工程量计算书.xlsx。v6.4: 0量待提取项不再静默剔除 —
    追加到表尾独立区域(标'待提取'), 交付可见, 不编造量。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = '工程量计算表'
    hf = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    pf = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    hft = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF'); df = Font(name='微软雅黑', size=10)
    bd = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    ws.cell(row=1,column=1,value='工程量计算书').font = Font(name='微软雅黑', bold=True, size=16)
    ws.cell(row=1,column=1).alignment = Alignment(horizontal='center'); ws.merge_cells('A1:J1')
    headers = ['序号','分项名称','单位','工程量','计算式','定额编号','规则来源','备注','计算来源','专业','部位','依据']
    for i,h in enumerate(headers,1):
        c=ws.cell(row=3,column=i,value=h); c.font=hft; c.fill=hf; c.alignment=Alignment(horizontal='center'); c.border=bd
    r = 4
    for ri,item in enumerate(results):
        vals=[ri+1,item.get('分项名称',''),item.get('单位',''),item.get('工程量',0),item.get('计算式',''),item.get('定额编号',''),item.get('规则来源',''),item.get('备注',''),item.get('_engine',''),item.get('专业',''),item.get('部位',''),item.get('依据','')]
        for ci,v in enumerate(vals,1):
            ws.cell(row=r,column=ci,value=v).font=df; ws.cell(row=r,column=ci).border=bd; ws.cell(row=r,column=ci).alignment=Alignment(horizontal='center', wrap_text=True)
        r += 1
    # v6.4: 待提取分项(0量占位) — 独立区域, 交付可见
    if pending:
        r += 1
        ws.cell(row=r,column=2,value='▼ 待提取分项（0量占位，图纸无面积/数量证据，需人工补量后重算）').font = Font(name='微软雅黑', bold=True, size=10, color='C00000')
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=10)
        r += 1
        for ri,item in enumerate(pending,1):
            vals=[ri,item.get('分项名称',''),item.get('单位',''),item.get('工程量',0),item.get('计算式',''),item.get('定额编号',''),item.get('规则来源',''),item.get('备注',''),item.get('_engine',''),item.get('专业','')]
            for ci,v in enumerate(vals,1):
                c=ws.cell(row=r,column=ci,value=v); c.font=df; c.border=bd; c.fill=pf
                c.alignment=Alignment(horizontal='center', wrap_text=True)
            r += 1
    for c,w in [('A',6),('B',26),('C',8),('D',10),('E',32),('F',12),('G',18),('H',18),('I',10),('J',10),('K',20),('L',40)]: ws.column_dimensions[c].width = w
    wb.save(path)


def split_pending_items(results):
    """v5.12 卡口强化: 0 量占位项从算量结果中剔除, 分流到待提取清单。
    占位项 = 工程量<=0 且 (数据来源=='待提取' 或 计算式含 '待CAD提取/待输入/待CAD图块')。
    返回 (有效结果, 待提取项)。step4 的拦截作为双保险保留。
    """
    valid, pending = [], []
    for it in results:
        qty = it.get('工程量', 0) or 0
        calc = str(it.get('计算式', ''))
        src = it.get('数据来源', '')
        is_placeholder = qty <= 0 and (
            src == '待提取' or
            '待CAD提取' in calc or '待输入' in calc or '待CAD图块' in calc or
            it.get('房间分区') is True  # v5.15 精装房间分区项(量待几何分区)
        )
        if is_placeholder:
            it['数据来源'] = '待提取'
            pending.append(it)
        else:
            valid.append(it)
    return valid, pending


def run(input_path, output_dir):
    print('='*50); print('Step 3: 算量 — v3.1 增强版'); print('='*50)
    with open(input_path, 'r', encoding='utf-8') as f:
        drawing_data = json.load(f)
    specialty = drawing_data.get('专业类型', '未知')
    print(f'  专业: {specialty}')
    results = calculate(drawing_data)
    # v5.12: 0 量占位项卡口 — 不进算量结果, 独立交付待提取清单
    results, pending_items = split_pending_items(results)
    report = quality_report(drawing_data, results)
    print(f'  总计: {len(results)} 个分项 | 待提取(0量占位): {len(pending_items)} 项 | 质量分: {report["质量分"]} | 警告: {report["警告数量"]}条')
    if pending_items:
        print(f'  ⚠ 已剔除 {len(pending_items)} 个 0 量占位项 → 待提取清单.json')
    xlsx_path = os.path.join(output_dir, '工程量计算书.xlsx')
    export_excel(results, xlsx_path, pending_items)
    print(f'  输出: {xlsx_path}')
    json_path = os.path.join(output_dir, '算量结果.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    with open(os.path.join(output_dir, '待提取清单.json'), 'w', encoding='utf-8') as f:
        json.dump(pending_items, f, ensure_ascii=False, indent=2)
    with open(os.path.join(output_dir, '算量质量报告.json'), 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    # v6.5 C1增强: 施工范围掩码(公共函数) — 设计内容 + 不含项 裁剪
    _apply_scope_mask(drawing_data, results)
    return results


if __name__ == '__main__':
    for sp in ['房屋建筑与装饰工程','市政工程','安装工程','园林绿化工程','钢结构工程']:
        test = {'专业类型': sp, '面积区域':[{'面积_m2':500}],'构造层':[{'名称':'混凝土','厚度_mm':120}], '线性构件':[{'名称':'管道','长度_m':30}], '施工说明':['挖土方','混凝土浇筑','钢筋','抹灰','防水','安装管道','电缆']}
        print(f'\n{sp}:')
        for item in calculate(test):
            print(f'  {item["分项名称"]}: {item["工程量"]} {item["单位"]} ({item.get("_engine","")})')
